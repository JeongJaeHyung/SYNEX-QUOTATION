# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "httpx",
#     "openpyxl",
# ]
# ///

import asyncio
import json
import os
from datetime import datetime

import httpx
from openpyxl import load_workbook

# --- Configuration ---
API_BASE_URL = "http://localhost:8000/api/v1"
MACHINE_API_URL = f"{API_BASE_URL}/quotation/machine/"
MACHINE_SEARCH_URL = f"{API_BASE_URL}/quotation/machine/search"
PARTS_LIST_URL = f"{API_BASE_URL}/parts"
PARTS_CREATE_URL = f"{API_BASE_URL}/parts/"
MAKER_CREATE_URL = f"{API_BASE_URL}/maker/"

EXCEL_FILE_PATH = "tools/data.xlsx"
TEMPLATE_SHEET_INDICES = list(range(3, 9))

MACHINE_NAME_PREFIX = "[TEMPLATE] "
TEMPLATE_CREATOR = "TEMPLATE"
AUTO_CREATE_MISSING_PARTS = True
UPSERT_TEMPLATES = True

# SQLite 안정성을 위한 동시성 제한
# 조희(GET)는 여러 개 가능하지만, 쓰기(POST/PUT) 시 경합 방지를 위해 1로 설정 권장
WRITE_SEMAPHORE = asyncio.Semaphore(1)

SUMMARY_MAJORS = {"전장/제어부 집계"}
LABOR_MAJORS = {"인건비"}


# --- Utility Functions ---
def clean_value(v):
    cleaned = str(v).strip() if v is not None else ""
    return "" if cleaned in ("공백", "nan", "None") or not cleaned else cleaned


def clean_maker_name(v):
    cleaned = str(v).strip() if v is not None else ""
    if cleaned == "공백":
        return " "
    return "" if cleaned in ("nan", "None") or not cleaned else cleaned


def normalize_key(v: str) -> str:
    return " ".join(str(v).strip().split()).casefold() if v else ""


def parse_int(v, default=0) -> int:
    try:
        return int(float(str(v).replace(",", "").strip()))
    except:
        return default


# --- Core Logic ---
async def fetch_parts_index(client: httpx.AsyncClient) -> dict:
    """DB의 모든 부품을 가져와 매칭용 인덱스를 생성합니다."""
    print("🔍 DB 부품 마스터 인덱싱 중...")
    index = {"by_maker_model": {}, "by_mm_empty_name": {}}
    skip, limit = 0, 1000
    while True:
        resp = await client.get(
            PARTS_LIST_URL, params={"skip": skip, "limit": limit}, timeout=30.0
        )
        data = resp.json()
        items = data.get("items", [])
        for item in items:
            m_name, p_name = (
                clean_maker_name(item.get("maker_name")),
                clean_value(item.get("name")),
            )
            major, minor = (
                clean_value(item.get("major_category")),
                clean_value(item.get("minor_category")),
            )
            payload = {
                "maker_id": item["maker_id"],
                "resources_id": item["id"],
                "solo_price": item.get("solo_price", 0),
            }

            # 매칭 키 생성
            m_k, p_k = normalize_key(m_name), normalize_key(p_name)
            if m_k and p_k:
                index["by_maker_model"][(m_k, p_k)] = payload
            if not p_k and major and minor:
                index["by_mm_empty_name"][
                    (normalize_key(major), normalize_key(minor))
                ] = payload

        skip += limit
        if skip >= data.get("total", 0):
            break
    return index


async def ensure_part(client: httpx.AsyncClient, row_data: dict, index: dict):
    """부품이 없으면 생성하고 인덱스에 추가합니다. (쓰기 락 적용)"""
    m_name, p_name = row_data["maker_name"], row_data["name"]
    major, minor = row_data["major_category"], row_data["minor_category"]

    # 1. 기존 인덱스 확인
    m_k, p_k = normalize_key(m_name), normalize_key(p_name)
    found = index["by_maker_model"].get((m_k, p_k))
    if not found and major and minor and not p_k:
        found = index["by_mm_empty_name"].get(
            (normalize_key(major), normalize_key(minor))
        )

    if found or not AUTO_CREATE_MISSING_PARTS:
        return found

    # 2. 신규 생성 (순차 처리)
    async with WRITE_SEMAPHORE:
        # 생성 전 한 번 더 확인 (중복 생성 방지)
        resp = await client.post(PARTS_CREATE_URL, json=row_data, timeout=15.0)
        if resp.status_code == 404:  # Maker 없음
            await client.post(MAKER_CREATE_URL, json={"name": m_name})
            resp = await client.post(PARTS_CREATE_URL, json=row_data)

        if 200 <= resp.status_code < 300:
            new_p = resp.json()
            res = {
                "maker_id": new_p["maker_id"],
                "resources_id": new_p["id"],
                "solo_price": new_p.get("solo_price", 0),
            }
            if m_k and p_k:
                index["by_maker_model"][(m_k, p_k)] = res
            return res
    return None


async def process_sheet(client: httpx.AsyncClient, ws, parts_index: dict):
    """개별 시트를 분석하여 Machine을 생성/갱신합니다."""
    title = ws.title
    print(f"📊 시트 처리 중: '{title}'")

    header = [
        str(c.value).strip() if c.value else ""
        for v in ws.iter_rows(min_row=1, max_row=1)
        for c in v
    ]
    h_map = {name: idx for idx, name in enumerate(header)}

    resources_payload = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean_value(row[h_map["모델명/규격"]])
        major = clean_value(row[h_map["Unit"]])
        if not name and major not in (SUMMARY_MAJORS | LABOR_MAJORS):
            continue
        if name == "-":
            continue

        row_info = {
            "maker_name": clean_maker_name(row[h_map["제조사"]]),
            "major_category": major,
            "minor_category": clean_value(row[h_map["품목"]]),
            "name": name or clean_value(row[h_map["품목"]]),  # 인건비용
            "unit": clean_value(row[h_map["단위"]])
            or ("M/D" if major in LABOR_MAJORS else "ea"),
            "solo_price": parse_int(row[h_map["단가"]]),
            "quantity": parse_int(row[h_map["수량"]]),
        }

        if row_info["quantity"] <= 0 and major not in LABOR_MAJORS:
            continue

        part = await ensure_part(client, row_info, parts_index)
        if part:
            resources_payload.append(
                {
                    "maker_id": part["maker_id"],
                    "resources_id": part["resources_id"],
                    "solo_price": row_info["solo_price"] or part["solo_price"],
                    "quantity": row_info["quantity"],
                    "display_major": major,
                    "display_minor": row_info["minor_category"],
                    "display_model_name": row_info["name"],
                }
            )

    # Machine 등록 (쓰기 락 적용)
    machine_name = f"{MACHINE_NAME_PREFIX}{title}"
    payload = {
        "name": machine_name,
        "creator": TEMPLATE_CREATOR,
        "resources": resources_payload,
    }

    async with WRITE_SEMAPHORE:
        # 기존 검색
        search_resp = await client.get(
            MACHINE_SEARCH_URL, params={"search": machine_name}
        )
        items = search_resp.json().get("items", [])
        existing = next((i for i in items if i["name"] == machine_name), None)

        if existing and UPSERT_TEMPLATES:
            await client.put(f"{MACHINE_API_URL}{existing['id']}", json=payload)
            print(f"✅ 갱신 완료: {machine_name}")
        else:
            await client.post(MACHINE_API_URL, json=payload)
            print(f"✅ 신규 등록: {machine_name}")


async def main():
    if not os.path.exists(EXCEL_FILE_PATH):
        return print("❌ 엑셀 파일 없음")
    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)

    async with httpx.AsyncClient(timeout=60.0) as client:
        parts_index = await fetch_parts_index(client)
        # 시트별로 비동기 작업 생성 (단, 내부 쓰기 작업은 세마포어로 보호됨)
        tasks = [
            process_sheet(client, wb.worksheets[i], parts_index)
            for i in TEMPLATE_SHEET_INDICES
            if i < len(wb.worksheets)
        ]
        await asyncio.gather(*tasks)


if __name__ == "__main__":
    asyncio.run(main())
