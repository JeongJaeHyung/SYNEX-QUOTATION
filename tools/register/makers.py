# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "httpx",
#     "openpyxl",
# ]
# ///

import asyncio
import os

import httpx
from openpyxl import load_workbook

# --- Configuration ---
API_URL = "http://localhost:8000/api/v1/maker"
EXCEL_FILE_PATH = "tools/data.xlsx"
MAKER_SHEET_NAME = 2  # 시트 이름 "maker"로 바꿔도 돼
CONCURRENCY_LIMIT = 50


def clean_value(value):
    """데이터 클렌징 및 '공백' 문자 처리"""
    if value is None:
        return ""
    cleaned = str(value).strip()
    if cleaned == "공백":
        return " "
    return "" if cleaned in ("nan", "None") or not cleaned else cleaned


def get_maker_data_from_excel(path, sheet_idx):
    if not os.path.exists(path):
        print(f"❌ 파일이 없어: {path}")
        return []

    # read_only=True로 로딩 속도 최적화
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[sheet_idx] if isinstance(sheet_idx, int) else wb[sheet_idx]
    except (IndexError, KeyError):
        print(f"❌ 시트 인덱스 {sheet_idx}를 찾을 수 없어.")
        return []

    print(f"📊 작업 중인 시트: {ws.title}")

    # 상단 10줄 이내에서 '회사코드'가 포함된 진짜 헤더 행 찾기
    rows_iter = ws.iter_rows(values_only=True)
    header = None
    h_map = {}

    for i, row in enumerate(rows_iter):
        row_str = [str(c).strip() if c is not None else "" for c in row]
        if "회사코드" in row_str and "회사명" in row_str:
            header = row_str
            h_map = {name: idx for idx, name in enumerate(header)}
            print(f"🔍 헤더 발견 (행 {i + 1}): {header}")
            break
        if i > 10:
            break

    if not header:
        print("❌ '회사코드' 또는 '회사명' 컬럼을 찾지 못했어. 엑셀을 확인해봐.")
        return []

    maker_list = []
    seen = set()
    # 헤더 다음 줄부터 끝까지 데이터 추출
    for row in rows_iter:
        m_id = clean_value(row[h_map["회사코드"]])
        name = clean_value(row[h_map["회사명"]])

        if m_id and name and m_id not in seen:
            seen.add(m_id)
            maker_list.append({"id": m_id, "name": name})

    return maker_list


async def post_maker(client, sem, data):
    """비동기 API 요청 및 중복(409) 처리"""
    async with sem:
        try:
            r = await client.post(API_URL, json=data, timeout=10.0)
            # 2xx 성공 혹은 409 이미 존재함은 정상으로 간주
            if r.status_code < 300 or r.status_code == 409:
                return "OK"
            return f"ERR({r.status_code})"
        except Exception as e:
            return f"FAIL({type(e).__name__})"


async def main():
    makers = get_maker_data_from_excel(EXCEL_FILE_PATH, MAKER_SHEET_NAME)
    if not makers:
        return

    print(f"🚀 {len(makers)}개 데이터 전송 시작 (동시 요청: {CONCURRENCY_LIMIT})...")

    sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
    async with httpx.AsyncClient() as client:
        # 모든 요청을 비동기로 생성 후 한꺼번에 실행
        tasks = [post_maker(client, sem, m) for m in makers]
        results = await asyncio.gather(*tasks)

    ok_count = results.count("OK")
    print("\n--- 전송 결과 ---")
    print(f"✅ 성공/중복: {ok_count}개")
    print(f"❌ 실패: {len(results) - ok_count}개")


if __name__ == "__main__":
    asyncio.run(main())
