# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "httpx",
#     "openpyxl",
# ]
# ///

import asyncio
import os

import httpx
from openpyxl import load_workbook

# --- Configuration ---
API_URL = "http://localhost:8000/api/v1/parts/"
EXCEL_FILE_PATH = "tools/data.xlsx"
PARTS_SHEET_NAME = 0
# 동시에 처리할 요청 수 (너무 높으면 SQLite Lock 발생)
CONCURRENCY_LIMIT = 1
# 요청 사이의 미세한 간격 (초)
REQUEST_DELAY = 0.03


def clean_value(value):
    if value is None:
        return ""
    cleaned = str(value).strip()
    if cleaned in ("공백", "nan", "None") or not cleaned:
        return ""
    return cleaned


def clean_maker_name(value):
    if value is None:
        return ""
    cleaned = str(value).strip()
    if cleaned == "공백":
        return " "
    if cleaned in ("nan", "None") or not cleaned:
        return ""
    return cleaned


async def get_existing_parts(client):
    """DB에 이미 등록된 부품 목록을 가져와서 (이름, 제조사명) 세트로 반환합니다."""
    try:
        resp = await client.get(API_URL, timeout=20.0)
        if resp.status_code == 200:
            data = resp.json()
            # 이름과 제조사명 조합으로 중복 체크 세트 생성
            return {(p["name"], p["maker_name"]) for p in data}
    except Exception as e:
        print(f"⚠️ 기존 데이터 조회 실패 (무시하고 진행): {e}")
    return set()


def get_parts_data_from_excel(path, sheet_idx):
    """엑셀에서 데이터를 추출하여 전송용 리스트로 만듭니다."""
    if not os.path.exists(path):
        return [], 0
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_idx] if isinstance(sheet_idx, int) else wb[sheet_idx]

    header_row_index = 1
    header = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in ws[header_row_index]
    ]
    h_map = {name: idx for idx, name in enumerate(header)}

    data_list = []
    skipped_count = 0

    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        raw_part_name = str(
            row[h_map["모델명/규격"]] if h_map["모델명/규격"] < len(row) else ""
        ).strip()
        if raw_part_name == "-":
            skipped_count += 1
            continue

        try:
            price_str = (
                str(row[h_map["단가"]] if h_map["단가"] < len(row) else "0")
                .replace(",", "")
                .strip()
            )
            solo_price = (
                int(float(price_str)) if price_str and price_str != "None" else 0
            )

            part_json = {
                "maker_name": clean_maker_name(row[h_map["제조사"]]),
                "major_category": clean_value(row[h_map["Unit"]]),
                "minor_category": clean_value(row[h_map["품목"]]),
                "name": clean_value(row[h_map["모델명/규격"]]),
                "unit": clean_value(row[h_map["단위"]]) or "ea",
                "solo_price": solo_price,
                "display_order": 0,  # 전송 전 임시값
                "ul": clean_value(row[h_map["UL"]]) == "O",
                "ce": clean_value(row[h_map["CE"]]) == "O",
                "kc": clean_value(row[h_map["KC"]]) == "O",
                "certification_etc": clean_value(row[h_map["기타"]]),
            }
            data_list.append(part_json)
        except Exception:
            skipped_count += 1

    return data_list, skipped_count


async def post_part(client, sem, data, index):
    """개별 부품을 전송하는 비동기 함수입니다."""
    async with sem:
        try:
            # 순서 보장을 위해 인덱스 재지정
            data["display_order"] = index
            # 속도 조절을 위한 미세 지연
            await asyncio.sleep(REQUEST_DELAY)

            r = await client.post(API_URL, json=data, timeout=15.0)
            if 200 <= r.status_code < 300:
                return "OK"
            if r.status_code == 409:
                return "EXIST"
            return f"ERR({r.status_code})"
        except Exception:
            return "FAIL"


async def main():
    async with httpx.AsyncClient() as client:
        # 1. 엑셀 데이터 파싱
        print("📊 엑셀 파일 읽는 중...")
        excel_parts, skipped = get_parts_data_from_excel(
            EXCEL_FILE_PATH, PARTS_SHEET_NAME
        )
        if not excel_parts:
            return print("❌ 등록할 부품이 없습니다.")

        # 2. DB에서 기존 부품 조회
        print("🔍 DB 중복 데이터 확인 중...")
        existing_set = await get_existing_parts(client)

        # 3. 이름/제조사 기준으로 필터링
        new_parts = []
        for p in excel_parts:
            if (p["name"], p["maker_name"]) in existing_set:
                continue
            new_parts.append(p)

        print(
            f"🚀 총 {len(excel_parts)}개 중 {len(new_parts)}개 신규 등록 시작 (중복 제외: {len(excel_parts) - len(new_parts)}개)..."
        )

        if not new_parts:
            return print("✅ 모든 부품이 이미 등록되어 있습니다.")

        # 4. 속도 조절하며 비동기 전송
        sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
        tasks = [post_part(client, sem, p, i) for i, p in enumerate(new_parts)]
        results = await asyncio.gather(*tasks)

    # 5. 결과 집계
    print("\n--- 부품 전송 완료 ---")
    print(f"✅ 신규 등록 성공: {results.count('OK')}개")
    print(f"⚠️ 이미 존재함: {results.count('EXIST')}개")
    print(f"❌ 실패: {len(results) - results.count('OK') - results.count('EXIST')}개")


if __name__ == "__main__":
    asyncio.run(main())
