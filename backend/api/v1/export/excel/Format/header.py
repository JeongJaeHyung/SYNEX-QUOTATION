# backend/api/v1/export/excel/format/header.py
from datetime import datetime
from io import BytesIO
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def number_to_korean(number: int) -> str:
    """
    JS의 numberToKorean과 동일한 로직
    일금 X만 X천 X백원 정 형식으로 변환
    """
    if number == 0:
        return "일금 영원 정"

    units = ["", "만", "억", "조", "경"]
    nums = ["영", "일", "이", "삼", "사", "오", "육", "칠", "팔", "구"]
    decimals = ["", "십", "백", "천"]

    str_num = str(number)
    result = ""
    unit_index = 0

    while len(str_num) > 0:
        chunk = str_num[-4:]
        str_num = str_num[:-4]
        chunk_res = ""

        for i, digit in enumerate(reversed(chunk)):
            d = int(digit)
            if d > 0:
                chunk_res = nums[d] + decimals[i] + chunk_res

        if chunk_res:
            result = chunk_res + units[unit_index] + result
        unit_index += 1

    return f"일금 {result}원 정"


def create_excel(data: dict[str, Any]) -> BytesIO:
    """
    견적서 갑지 Excel 생성

    Args:
        data: {
            'title': str,
            'client': str,
            'pic_name': str,
            'pic_position': str,
            'price': int,
            'resources': [
                {
                    'machine': str,
                    'name': str,
                    'spac': str,
                    'compare': int,
                    'unit': str,
                    'solo_price': int,
                    'subtotal': int,
                    'description': str
                }
            ],
            'description_1': str,  # 특이사항
            'description_2': str   # 납기/지불조건
        }

    Returns:
        BytesIO: Excel 파일 바이너리
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ========================================================================
    # 스타일 정의
    # ========================================================================
    thin = Side(style="thin")
    medium = Side(style="medium")

    # 폰트
    font_title = Font(name="맑은 고딕", size=33, bold=True)
    font_client = Font(name="맑은 고딕", size=15, bold=True)
    font_default = Font(name="맑은 고딕", size=11)
    font_bold_11 = Font(name="맑은 고딕", size=11, bold=True)
    font_bold_12 = Font(name="맑은 고딕", size=12, bold=True)
    font_header = Font(name="맑은 고딕", size=10, bold=True)
    font_data = Font(name="맑은 고딕", size=10)
    font_9 = Font(name="맑은 고딕", size=9)
    font_red_12 = Font(name="맑은 고딕", size=12, bold=True, color="FF0000")

    # 정렬
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 배경색
    fill_blue = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    fill_yellow = PatternFill(
        start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
    )
    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ========================================================================
    # 1. 컬럼 너비 설정 (JS와 동일)
    # ========================================================================
    widths = [4.0, 22.5, 13.0, 6.0, 5.5, 13.0, 13.375, 12.375, 9.0, 19.5, 5.625]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ========================================================================
    # Row 1: 공백 (높이 15)
    # ========================================================================
    ws.row_dimensions[1].height = 15

    # ========================================================================
    # Row 2: QUOTATION 타이틀
    # ========================================================================
    ws.merge_cells("A2:K2")
    cell = ws["A2"]
    cell.value = "QUOTATION"
    cell.font = font_title
    cell.alignment = align_center
    cell.border = Border(top=medium, left=medium, right=medium)

    # QUOTATION 전체 셀에 medium 테두리
    for col in range(1, 12):
        c = ws.cell(2, col)
        c.border = Border(
            top=medium,
            left=medium if col == 1 else thin,
            right=medium if col == 11 else thin,
        )
    ws.row_dimensions[2].height = 42.75

    # ========================================================================
    # Row 3: 날짜 및 공급자 정보 시작
    # ========================================================================
    ws.merge_cells("B3:D3")
    today = datetime.now()
    # Excel 날짜 시리얼 번호로 저장
    ws["B3"] = today
    ws["B3"].number_format = "YYYY-MM-DD"  # 날짜 포맷
    ws["B3"].font = font_default
    ws["B3"].alignment = align_center
    ws["B3"].border = Border(left=medium, right=medium)

    ws.merge_cells("F3:F7")
    ws["F3"] = "공 급 자"
    ws["F3"].font = font_bold_11
    ws["F3"].alignment = align_center
    ws["F3"].border = Border(top=medium, left=medium, bottom=medium)

    ws["G3"] = "견적 번호"
    ws["G3"].font = font_9
    ws["G3"].alignment = align_center
    ws["G3"].border = Border(top=medium, left=thin, right=thin)

    ws.merge_cells("H3:J3")
    ws["H3"] = data.get("quotation_number", "")
    ws["H3"].font = font_9
    ws["H3"].alignment = align_center
    ws["H3"].border = Border(top=medium)

    ws["K3"].border = Border(right=medium)

    # ========================================================================
    # Row 4-5: 고객사명 및 공급자 상호
    # ========================================================================
    ws.merge_cells("B4:D5")
    ws["B4"] = data.get("client")
    ws["B4"].font = font_client
    ws["B4"].alignment = align_center
    ws["B4"].border = Border(left=medium, right=medium)

    ws["G4"] = "상호(법인명)"
    ws["G4"].font = font_9
    ws["G4"].alignment = align_center
    ws["G4"].border = Border(left=thin, right=thin)

    # ⚠️ H4는 병합하지 않음! (I4에 "성  명" 라벨이 들어감)
    ws["H4"] = "(주)제이엘티"
    ws["H4"].font = font_9
    ws["H4"].alignment = align_center

    ws["I4"] = "성  명"
    ws["I4"].font = font_9
    ws["I4"].alignment = align_center

    ws["J4"] = "정현우"
    ws["J4"].font = font_9
    ws["J4"].alignment = align_center

    ws["K4"].border = Border(right=medium)

    # ========================================================================
    # Row 5: 사업장주소
    # ========================================================================
    ws["G5"] = "사업장주소"
    ws["G5"].font = font_9
    ws["G5"].alignment = align_center
    ws["G5"].border = Border(left=thin, right=thin)

    # ⚠️ H5:J5 병합 (K 포함 X)
    ws.merge_cells("H5:J5")
    ws["H5"] = "인천광역시 연수구 벤처로 100번길 12"
    ws["H5"].font = font_9
    ws["H5"].alignment = align_center

    ws["K5"].border = Border(right=medium)

    # ========================================================================
    # Row 6: 담당자 및 업태
    # ========================================================================
    ws.merge_cells("B6:D6")
    pic_name = data.get("pic_name")
    pic_position = data.get("pic_position")
    ws["B6"] = f"{pic_name} {pic_position}님 귀하"
    ws["B6"].font = font_default
    ws["B6"].alignment = align_center
    ws["B6"].border = Border(left=medium, right=medium)

    ws["G6"] = "업    태"
    ws["G6"].font = font_9
    ws["G6"].alignment = align_center
    ws["G6"].border = Border(left=thin, right=thin)

    # ⚠️ H6는 병합하지 않음! (I6에 "사업자번호" 라벨이 들어감)
    ws["H6"] = "제조,도,소매"
    ws["H6"].font = font_9
    ws["H6"].alignment = align_center

    ws["I6"] = "사업자번호"
    ws["I6"].font = font_9
    ws["I6"].alignment = align_center

    ws["J6"] = "362-81-01342"
    ws["J6"].font = font_9
    ws["J6"].alignment = align_center

    ws["K6"].border = Border(right=medium)

    # ========================================================================
    # Row 7: 인사말 및 연락처
    # ========================================================================
    ws.merge_cells("B7:D7")
    ws["B7"] = "아래와 같이 견적합니다."
    ws["B7"].font = font_default
    ws["B7"].alignment = align_center
    ws["B7"].border = Border(left=medium, right=medium, bottom=medium)

    ws["G7"] = "TEL"
    ws["G7"].font = font_9
    ws["G7"].alignment = align_center
    ws["G7"].border = Border(bottom=medium, left=thin, right=thin)

    # ⚠️ H7는 병합하지 않음! (I7에 "mail" 라벨이 들어감)
    ws["H7"] = "010-9187-9593"
    ws["H7"].font = font_9
    ws["H7"].alignment = align_center

    ws["I7"] = "mail"
    ws["I7"].font = font_9
    ws["I7"].alignment = align_center
    ws["I7"].border = Border(bottom=medium)

    ws["J7"] = "hwjeong@jltfa.co.kr"
    ws["J7"].font = font_9
    ws["J7"].alignment = align_center
    ws["J7"].border = Border(bottom=medium)

    ws["K7"].border = Border(bottom=medium, right=medium)

    # ========================================================================
    # Row 8-9: 제목 및 견적금액
    # ========================================================================
    ws.merge_cells("A8:A9")
    ws.merge_cells("B8:D9")
    title = data.get("title", "HMC 차세대 배터리 라인")
    ws["B8"] = f"제목 : {title}"
    ws["B8"].font = font_bold_12
    ws["B8"].alignment = align_wrap

    total_price = data.get("price", 0)

    # H8: 견적금액 :
    ws["H8"] = "견적금액 :"
    ws["H8"].font = font_bold_11
    ws["H8"].alignment = align_center

    # I8:K8 병합 - 한글 금액
    ws.merge_cells("I8:K8")
    ws["I8"] = number_to_korean(total_price)
    ws["I8"].font = font_bold_11
    ws["I8"].alignment = align_left

    # I9:K9 병합 - ₩ 금액
    ws.merge_cells("I9:K9")
    ws["I9"] = f"₩{total_price:,} (VAT별도)"
    ws["I9"].font = font_bold_11
    ws["I9"].alignment = align_right

    # ========================================================================
    # Row 10: 문서 타이틀
    # ========================================================================
    ws.merge_cells("A10:K10")
    ws["A10"] = title
    ws["A10"].font = font_bold_11
    ws["A10"].alignment = align_center
    ws["A10"].fill = fill_blue

    for col in range(1, 12):
        c = ws.cell(10, col)
        c.border = Border(
            top=medium,
            bottom=medium,
            left=medium,
            right=medium,
        )
    ws.row_dimensions[10].height = 17.25

    # ========================================================================
    # Row 11: 테이블 헤더
    # ========================================================================
    ws.merge_cells("H11:I11")  # 공급가액
    ws.merge_cells("J11:K11")  # 비고

    headers = [
        "No",
        "장비명",
        "품     명",
        "규 격",
        "수량",
        "단위",
        "단 가",
        "공급가액",
        "",
        "비 고",
        "",
    ]
    for i, h in enumerate(headers, 1):
        if i in [9, 11]:  # 병합된 셀 건너뛰기
            continue

        cell = ws.cell(11, i)
        cell.value = h
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_yellow
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # ========================================================================
    # 동적 데이터 행 (Row 12+)
    # ========================================================================
    curr_row = 12
    resources = data.get("resources", [])

    for idx, item in enumerate(resources, 1):
        # No
        ws.cell(curr_row, 1, idx)
        # 장비명
        ws.cell(curr_row, 2, item.get("machine", ""))
        # 품명
        ws.cell(curr_row, 3, item.get("name", ""))
        # 규격
        ws.cell(curr_row, 4, item.get("spac", ""))
        # 수량
        ws.cell(curr_row, 5, item.get("compare", 1))
        # 단위
        ws.cell(curr_row, 6, item.get("unit", "원"))
        # 단가
        cell_price = ws.cell(curr_row, 7, item.get("solo_price", 0))
        cell_price.number_format = "#,##0"

        # 공급가액 (H:I 병합)
        ws.merge_cells(f"H{curr_row}:I{curr_row}")
        subtotal = item.get("subtotal", 0)
        cell_subtotal = ws.cell(curr_row, 8, subtotal)
        cell_subtotal.number_format = "#,##0"

        # 비고 (J:K 병합)
        ws.merge_cells(f"J{curr_row}:K{curr_row}")
        ws.cell(curr_row, 10, item.get("description", ""))

        # 스타일 적용
        for col in range(1, 12):
            c = ws.cell(curr_row, col)
            c.font = font_data
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            # 정렬
            if col <= 6:
                c.alignment = align_center
            elif col == 7:
                c.alignment = align_right
            elif col in [8, 9]:
                c.alignment = align_right
            else:
                c.alignment = align_left

        curr_row += 1

    # ========================================================================
    # 견적 총 합계 (동적 위치)
    # ========================================================================
    summary_row = curr_row
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    cell_summary = ws.cell(summary_row, 1, "견적 총 합계")
    cell_summary.font = font_bold_11
    cell_summary.alignment = align_center
    cell_summary.border = Border(top=medium, bottom=medium, left=medium)

    # 총 수량
    total_qty = sum(r.get("compare", 1) for r in resources)
    ws.cell(summary_row, 5, total_qty)
    ws.cell(summary_row, 5).alignment = align_center
    ws.cell(summary_row, 5).border = Border(top=medium, bottom=medium)
    # Set
    ws.cell(summary_row, 6, "Set")
    ws.cell(summary_row, 6).alignment = align_center
    ws.cell(summary_row, 6).border = Border(top=medium, bottom=medium)
    # 단가
    ws.cell(summary_row, 7, "")
    ws.cell(summary_row, 7).alignment = align_center
    ws.cell(summary_row, 7).border = Border(top=medium, bottom=medium)


    # 합계 (H:I 병합)
    ws.merge_cells(f"H{summary_row}:I{summary_row}")
    total_subtotal = sum(r.get("subtotal", 0) for r in resources)
    cell_total = ws.cell(summary_row, 8, total_subtotal)
    cell_total.font = font_bold_11
    cell_total.number_format = "#,##0"
    cell_total.alignment = align_right
    cell_total.border = Border(top=medium, bottom=medium)

    ws.row_dimensions[summary_row].height = 17.25
    curr_row += 1

    # 빈 셀들에 테두리 적용
    ws.merge_cells(f"J{summary_row}:K{summary_row}")
    ws.cell(summary_row, 10).border = Border(top=medium, right=medium, bottom=medium)


    # ========================================================================
    # 비고 (특이사항) - 동적 위치
    # ========================================================================
    ws.merge_cells(f"A{curr_row}:K{curr_row}")
    cell_note_title = ws.cell(curr_row, 1, "비    고 ( 특이사항 )")
    cell_note_title.font = font_bold_11
    cell_note_title.alignment = align_center
    cell_note_title.border = Border(top=medium, left=medium, right=medium, bottom=medium)
    curr_row += 1

    ws.merge_cells(f"A{curr_row}:K{curr_row}")
    special_remarks = data.get("description_1", "1. 2개라인 기준의 견적서 입니다.")
    ws.cell(curr_row, 1, special_remarks)
    ws.cell(curr_row, 1).font = font_data
    ws.cell(curr_row, 1).alignment = align_left
    ws.cell(curr_row, 1).border = Border(left=medium, right=medium)
    curr_row += 1

    # ========================================================================
    # 빈 행 (13개 고정)
    # ========================================================================
    for _ in range(13):
        ws.merge_cells(f"A{curr_row}:K{curr_row}")
        curr_row += 1

    # ========================================================================
    # Total / Best nego Total
    # ========================================================================
    # Total 행
    ws.merge_cells(f"A{curr_row}:I{curr_row}")
    cell_total_label = ws.cell(curr_row, 1, "Total")
    cell_total_label.font = font_bold_12
    cell_total_label.alignment = align_center
    cell_total_label.border = Border(top=medium, left=medium, bottom=thin)

    # Total 행 전체 테두리
    for col in range(2, 10):
        ws.cell(curr_row, col).border = Border(top=medium, bottom=thin)

    ws.merge_cells(f"J{curr_row}:K{curr_row}")
    cell_total_value = ws.cell(curr_row, 10, total_price)
    cell_total_value.font = font_bold_12
    cell_total_value.number_format = "#,##0"
    cell_total_value.alignment = align_right
    cell_total_value.border = Border(top=medium, bottom=thin)

    ws.cell(curr_row, 11).border = Border(top=medium, right=medium, bottom=thin)
    curr_row += 1

    # Best nego Total 행
    ws.merge_cells(f"A{curr_row}:I{curr_row}")
    cell_nego_label = ws.cell(curr_row, 1, "Best nego Total")
    cell_nego_label.font = font_red_12
    cell_nego_label.alignment = align_center
    cell_nego_label.border = Border(top=thin, left=medium)

    # Best nego 행 전체 테두리
    for col in range(2, 10):
        ws.cell(curr_row, col).border = Border(top=thin)

    ws.merge_cells(f"J{curr_row}:K{curr_row}")
    cell_nego_value = ws.cell(curr_row, 10, total_price)
    cell_nego_value.font = font_red_12
    cell_nego_value.number_format = "#,##0"
    cell_nego_value.alignment = align_right
    cell_nego_value.border = Border(top=thin)

    ws.cell(curr_row, 11).border = Border(top=thin, right=medium)
    ws.row_dimensions[curr_row].height = 17.25
    curr_row += 1

    # ========================================================================
    # 비고 (납기/지불조건 등)
    # ========================================================================

    # 비고 라벨 (A:B, 5행 병합)
    ws.merge_cells(f"A{curr_row}:B{curr_row + 4}")
    cell_remarks_label = ws.cell(curr_row, 1, "비 고")
    cell_remarks_label.font = font_bold_11
    cell_remarks_label.alignment = align_center
    cell_remarks_label.border = Border(top=medium, left=medium, bottom=medium)

    ws.cell(curr_row, 2).border = Border(top=medium, right=medium, bottom=medium)

    # 비고 좌측 셀 세로 병합 테두리
    for r in range(curr_row + 1, curr_row + 5):
        ws.cell(r, 1).border = Border(left=medium)
        ws.cell(r, 2).border = Border(right=medium)

    ws.cell(curr_row + 4, 1).border = Border(left=medium, bottom=medium)
    ws.cell(curr_row + 4, 2).border = Border(right=medium, bottom=medium)

    # 비고 내용
    description_2 = data.get(
        "description_2",
        "- 납기 : 협의사항\n- 지불조건 : 선급금 30%, 중도금 50%, 잔금 20%\n- 기타 : 견적유효기간 10 일",  # noqa: E501
    )
    remarks_lines = description_2.split("\n")[:5]
    while len(remarks_lines) < 5:
        remarks_lines.append("")

    for i in range(5):
        ws.merge_cells(f"C{curr_row}:K{curr_row}")
        ws.cell(curr_row, 3, remarks_lines[i])
        ws.cell(curr_row, 3).font = font_data
        ws.cell(curr_row, 3).alignment = align_left

        # 비고 우측 테두리
        ws.cell(curr_row, 3).border = Border(
            top=medium if i == 0 else None, left=medium
        )
        ws.cell(curr_row, 11).border = Border(
            top=medium if i == 0 else None,
            right=medium,
            bottom=medium if i == 4 else None,
        )

        # 중간 컬럼들 상하 테두리만
        for col in range(4, 11):
            ws.cell(curr_row, col).border = Border(
                top=medium if i == 0 else None, bottom=medium if i == 4 else None
            )

        curr_row += 1

    # ========================================================================
    # 두꺼운 테두리 적용
    # ========================================================================
    # A2:K(마지막 행) 전체에 두꺼운 외곽 테두리
    last_row = curr_row - 1
    for row in range(2, last_row + 1):
        for col in range(1, 12):  # A(1) ~ K(11)
            cell = ws.cell(row, col)
            current_border = cell.border

            # 기존 테두리 유지하면서 외곽만 두껍게
            new_border = Border(
                left=medium if col == 1 else current_border.left,
                right=medium if col == 11 else current_border.right,
                top=medium if row == 2 else current_border.top,
                bottom=medium if row == last_row else current_border.bottom
            )
            cell.border = new_border

    # B3:D7에 두꺼운 테두리
    for row in range(3, 8):  # 3~7
        for col in range(2, 5):  # B(2) ~ D(4)
            cell = ws.cell(row, col)
            current_border = cell.border

            new_border = Border(
                left=medium if col == 2 else current_border.left,
                right=medium if col == 4 else current_border.right,
                top=medium if row == 3 else current_border.top,
                bottom=medium if row == 7 else current_border.bottom
            )
            cell.border = new_border

    # F3:J7에 두꺼운 테두리
    for row in range(3, 8):  # 3~7
        for col in range(6, 11):  # F(6) ~ J(10)
            cell = ws.cell(row, col)
            current_border = cell.border

            new_border = Border(
                left=medium if col == 6 else current_border.left,
                right=medium if col == 10 else current_border.right,
                top=medium if row == 3 else current_border.top,
                bottom=medium if row == 7 else current_border.bottom
            )
            cell.border = new_border

    # K3, K7의 상단/하단 두꺼운 테두리 제거 (기존 테두리 유지)
    k3_cell = ws.cell(3, 11)
    k3_cell.border = Border(top=medium, right=medium)  # 상단은 A2:K2 전체 테두리와 일치

    k7_cell = ws.cell(7, 11)
    k7_cell.border = Border(right=medium)  # 하단 두꺼운 테두리 제거

    # ========================================================================
    # Excel 파일 생성 및 반환
    # ========================================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
