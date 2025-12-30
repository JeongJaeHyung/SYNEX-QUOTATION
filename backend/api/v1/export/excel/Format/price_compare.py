# backend/api/v1/export/excel/format/price_compare.py
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def group_by_major_then_machine(
    resources: list[dict[str, Any]],
) -> dict[str, dict[str, list]]:
    """
    major(항목) -> machine_name(장비명)으로 2단계 그룹화
    """
    groups = {}

    for item in resources:
        major = item.get("major", "기타")
        machine = item.get("machine_name", "미분류")

        if major not in groups:
            groups[major] = {}
        if machine not in groups[major]:
            groups[major][machine] = []

        groups[major][machine].append(item)

    return groups


def create_excel(data: dict[str, Any]) -> BytesIO:
    """
    내정가견적비교서 Excel 파일 생성

    Args:
        data: {
            "id": str,
            "creator": str,
            "description": str,
            "created_at": datetime,
            "updated_at": datetime,
            "resources": [
                {
                    "machine_id": str,
                    "machine_name": str,
                    "major": str,
                    "minor": str,
                    "cost_solo_price": int,
                    "cost_unit": str,
                    "cost_compare": int,
                    "quotation_solo_price": int,
                    "quotation_unit": str,
                    "quotation_compare": int,
                    "upper": float,
                    "description": str
                }
            ]
        }

    Returns:
        BytesIO: Excel 파일 바이너리
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "내정가비교"

    # ========================================================================
    # 스타일 정의
    # ========================================================================
    thin = Side(style="thin")
    thick = Side(style="thick")
    double = Side(style="double")

    # 폰트
    font_default = Font(name="맑은 고딕", size=10)
    font_bold = Font(name="맑은 고딕", size=10, bold=True)
    font_title = Font(name="맑은 고딕", size=18, bold=True)
    font_red = Font(name="맑은 고딕", size=10, color="DC2626")

    # 정렬
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 테두리
    border_base = Border(top=thin, bottom=thin, left=thin, right=thin)
    border_subtotal = Border(top=thick, bottom=thick, left=thin, right=thin)
    border_total = Border(top=double, bottom=thick, left=thin, right=thin)

    # 배경색
    fill_header = PatternFill(
        start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
    )  # 하늘색
    fill_subtotal = PatternFill(
        start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"
    )  # 노란색
    fill_total = PatternFill(
        start_color="FDE68A", end_color="FDE68A", fill_type="solid"
    )  # 진한 노란색
    fill_note_head = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )  # 회색

    # ========================================================================
    # 컬럼 너비 설정 (13개 컬럼)
    # ========================================================================
    widths = [12, 18, 25, 6, 6, 14, 15, 6, 6, 14, 8, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ========================================================================
    # Row 1: 타이틀
    # ========================================================================
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = "내 정 가 / 견 적 가 비 교 서"
    title_cell.font = font_title
    title_cell.alignment = align_center

    # ========================================================================
    # Row 2-3: 테이블 헤더 (2단 구조)
    # ========================================================================
    # Row 2: 메인 헤더
    headers_main = [
        ("A2", "항목", 2),  # rowspan=2
        ("B2", "장비명", 2),
        ("C2", "구분", 2),
        ("D2", "내정가", 4),  # colspan=4 (D2:G2)
        ("H2", "견적가", 5),  # colspan=5 (H2:L2)
        ("M2", "비고", 2),
    ]

    for cell_addr, value, _span in headers_main:
        cell = ws[cell_addr]
        cell.value = value
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_base

    # 병합
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws.merge_cells("D2:G2")
    ws.merge_cells("H2:L2")
    ws.merge_cells("M2:M3")

    # Row 3: 서브 헤더
    sub_headers = [
        ("D3", "수량"),
        ("E3", "단위"),
        ("F3", "단가"),
        ("G3", "금액"),
        ("H3", "수량"),
        ("I3", "단위"),
        ("J3", "단가"),
        ("K3", "상승률"),
        ("L3", "금액"),
    ]

    for cell_addr, value in sub_headers:
        cell = ws[cell_addr]
        cell.value = value
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_base

    # ========================================================================
    # 데이터 렌더링
    # ========================================================================
    resources = data.get("resources", [])
    groups = group_by_major_then_machine(resources)

    # major 순서 정의
    major_order = ["자재비", "인건비", "출장경비", "관리비"]

    # major 정렬 (major_order 순서, 나머지는 뒤로)
    sorted_majors = sorted(
        groups.keys(), key=lambda x: major_order.index(x) if x in major_order else 99
    )

    curr_row = 4
    grand_cost_total = 0
    grand_quote_total = 0

    for major in sorted_majors:
        machine_groups = groups[major]
        start_major_row = curr_row
        major_cost = 0
        major_quote = 0

        for machine_name, items in machine_groups.items():
            start_machine_row = curr_row

            for idx, item in enumerate(items):
                # 금액 계산
                cost_amt = (item.get("cost_compare", 0) or 0) * (
                    item.get("cost_solo_price", 0) or 0
                )
                quote_amt = (item.get("quotation_compare", 0) or 0) * (
                    item.get("quotation_solo_price", 0) or 0
                )
                major_cost += cost_amt
                major_quote += quote_amt

                # 항목 (major) - 첫 행에만
                if idx == 0 and start_major_row == curr_row:
                    cell = ws.cell(curr_row, 1, major)
                else:
                    cell = ws.cell(curr_row, 1, "")
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 장비명 (machine_name) - 각 장비의 첫 행에만
                if idx == 0:
                    cell = ws.cell(curr_row, 2, machine_name)
                else:
                    cell = ws.cell(curr_row, 2, "")
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 구분 (minor)
                cell = ws.cell(curr_row, 3, item.get("minor", ""))
                cell.font = font_default
                cell.alignment = align_left
                cell.border = border_base

                # 내정가 - 수량
                cell = ws.cell(curr_row, 4, item.get("cost_compare", 0))
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 내정가 - 단위
                cell = ws.cell(curr_row, 5, "식")
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 내정가 - 단가
                cell = ws.cell(curr_row, 6, item.get("cost_solo_price", 0))
                cell.font = font_default
                cell.alignment = align_right
                cell.border = border_base
                cell.number_format = "#,##0"

                # 내정가 - 금액
                cell = ws.cell(curr_row, 7, cost_amt)
                cell.font = font_default
                cell.alignment = align_right
                cell.border = border_base
                cell.number_format = "#,##0"

                # 견적가 - 수량
                cell = ws.cell(curr_row, 8, item.get("quotation_compare", 0))
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 견적가 - 단위
                cell = ws.cell(curr_row, 9, "식")
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 견적가 - 단가
                cell = ws.cell(curr_row, 10, item.get("quotation_solo_price", 0))
                cell.font = font_default
                cell.alignment = align_right
                cell.border = border_base
                cell.number_format = "#,##0"

                # 견적가 - 상승률
                upper = item.get("upper", 0) or 0
                cell = ws.cell(curr_row, 11, f"{upper}%")
                cell.font = font_red
                cell.alignment = align_center
                cell.border = border_base

                # 견적가 - 금액
                cell = ws.cell(curr_row, 12, quote_amt)
                cell.font = font_bold
                cell.alignment = align_right
                cell.border = border_base
                cell.number_format = "#,##0"

                # 비고
                cell = ws.cell(curr_row, 13, item.get("description", ""))
                cell.font = font_default
                cell.alignment = align_left
                cell.border = border_base

                curr_row += 1

            # 장비명 셀 병합 (동일 장비의 여러 행)
            if len(items) > 1:
                ws.merge_cells(f"B{start_machine_row}:B{curr_row - 1}")

        # major 소계 행
        ws.merge_cells(f"A{curr_row}:F{curr_row}")
        subtotal_cell = ws.cell(curr_row, 1)
        subtotal_cell.value = f"{major} 소계"
        subtotal_cell.font = font_bold
        subtotal_cell.alignment = align_center
        subtotal_cell.fill = fill_subtotal
        subtotal_cell.border = Border(top=thick, bottom=thick, left=thin, right=thin)

        # A:F 병합된 셀의 테두리 적용
        for col in [2, 3, 4, 5, 6]:
            cell = ws.cell(curr_row, col)
            cell.fill = fill_subtotal
            cell.border = Border(top=thick, bottom=thick, left=thin, right=thin)

        # 빈 셀들 (H-K)
        for col in [8, 9, 10, 11]:
            cell = ws.cell(curr_row, col)
            cell.fill = fill_subtotal
            cell.border = border_subtotal

        # 내정가 소계 (G)
        cell = ws.cell(curr_row, 7, major_cost)
        cell.font = font_bold
        cell.alignment = align_right
        cell.fill = fill_subtotal
        cell.border = border_subtotal
        cell.number_format = "#,##0"

        # 견적가 소계 (L)
        cell = ws.cell(curr_row, 12, major_quote)
        cell.font = font_bold
        cell.alignment = align_right
        cell.fill = fill_subtotal
        cell.border = border_subtotal
        cell.number_format = "#,##0"

        # 차액 (M)
        cell = ws.cell(curr_row, 13, major_quote - major_cost)
        cell.font = font_bold
        cell.alignment = align_right
        cell.fill = fill_subtotal
        cell.border = border_subtotal
        cell.number_format = "#,##0"

        # 항목 셀 병합 (세로) 및 가운데 맞춤
        ws.merge_cells(f"A{start_major_row}:A{curr_row}")
        # 병합된 셀의 첫 번째 셀에 가운데 맞춤 적용
        major_cell = ws.cell(start_major_row, 1)
        major_cell.alignment = align_center

        grand_cost_total += major_cost
        grand_quote_total += major_quote
        curr_row += 1

    # ========================================================================
    # TOTAL 행
    # ========================================================================
    ws.merge_cells(f"A{curr_row}:F{curr_row}")
    total_cell = ws.cell(curr_row, 1)
    total_cell.value = "TOTAL"
    total_cell.font = font_bold
    total_cell.alignment = align_center
    total_cell.fill = fill_total
    total_cell.border = border_total

    # 빈 셀들
    for col in range(2, 7):
        cell = ws.cell(curr_row, col)
        cell.fill = fill_total
        cell.border = border_total

    # 내정가 합계 (G)
    cell = ws.cell(curr_row, 7, grand_cost_total)
    cell.font = font_bold
    cell.alignment = align_right
    cell.fill = fill_total
    cell.border = border_total
    cell.number_format = "#,##0"

    # 빈 셀들 (H-K)
    for col in range(8, 12):
        cell = ws.cell(curr_row, col)
        cell.fill = fill_total
        cell.border = border_total

    # 견적가 합계 (L)
    cell = ws.cell(curr_row, 12, grand_quote_total)
    cell.font = font_bold
    cell.alignment = align_right
    cell.fill = fill_total
    cell.border = border_total
    cell.number_format = "#,##0"

    # 차액 (M)
    cell = ws.cell(curr_row, 13, grand_quote_total - grand_cost_total)
    cell.font = font_bold
    cell.alignment = align_right
    cell.fill = fill_total
    cell.border = border_total
    cell.number_format = "#,##0"

    curr_row += 1

    # ========================================================================
    # 빈 행
    # ========================================================================
    curr_row += 1

    # ========================================================================
    # 비고 섹션
    # ========================================================================
    ws.merge_cells(f"A{curr_row}:M{curr_row}")
    note_head_cell = ws.cell(curr_row, 1)
    note_head_cell.value = "비 고 ( Note )"
    note_head_cell.font = font_bold
    note_head_cell.alignment = align_center
    note_head_cell.fill = fill_note_head
    note_head_cell.border = border_base
    curr_row += 1

    # 비고 내용 (4행 병합)
    ws.merge_cells(f"A{curr_row}:M{curr_row + 3}")
    note_content = data.get("description", "특이사항 없음")
    note_cell = ws.cell(curr_row, 1)
    note_cell.value = note_content
    note_cell.font = font_default
    note_cell.alignment = align_top_wrap
    note_cell.border = border_base

    # 병합된 셀의 테두리 적용
    for row in range(curr_row, curr_row + 4):
        for col in range(1, 14):
            ws.cell(row, col).border = border_base

    # ========================================================================
    # Excel 파일 생성 및 반환
    # ========================================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
