# backend/api/v1/export/excel/format/detailed.py
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 출력용 타이틀 매핑 (JavaScript의 MAJOR_DISPLAY_MAP과 동일)
MAJOR_DISPLAY_MAP = {
    "자재비": "자재비 상세 내역",
    "인건비": "인건비 상세 내역",
    "출장경비": "경비 상세 내역_국내",
    "관리비": "관리비 상세 내역",
}


def group_by_major_then_machine(
    resources: list[dict[str, Any]],
) -> dict[str, dict[str, list]]:
    """
    major(항목) -> machine_name(장비명)으로 2단계 그룹화
    공백 제거하여 정규화 (출장 경비 → 출장경비)
    """
    groups = {}

    for item in resources:
        raw_major = item.get("major", "기타")
        # 공백 제거하여 정규화
        major = raw_major.replace(" ", "")
        machine = item.get("machine_name", "미분류")

        if major not in groups:
            groups[major] = {}
        if machine not in groups[major]:
            groups[major][machine] = []

        groups[major][machine].append(item)

    return groups


def create_excel(data: dict[str, Any]) -> BytesIO:
    """
    견적서(을지) 상세 견적서 Excel 파일 생성

    Args:
        data: {
            "id": str,
            "title": str,
            "creator": str,
            "description": str,
            "created_at": datetime,
            "updated_at": datetime,
            "resources": [
                {
                    "machine_name": str,
                    "major": str,
                    "minor": str,
                    "spec": str,
                    "compare": int,
                    "unit": str,
                    "solo_price": int,
                    "description": str
                }
            ]
        }

    Returns:
        BytesIO: Excel 파일 바이너리
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "을지"

    # ========================================================================
    # 스타일 정의
    # ========================================================================
    thin = Side(style="thin")
    thick = Side(style="thick")

    # 기본 스타일
    font_default = Font(name="맑은 고딕", size=10)
    font_bold = Font(name="맑은 고딕", size=10, bold=True)
    font_title = Font(name="맑은 고딕", size=18, bold=True)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    border_base = Border(top=thin, bottom=thin, left=thin, right=thin)
    border_subtotal = Border(top=thick, bottom=thick, left=thin, right=thin)

    # 배경색
    fill_header = PatternFill(
        start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"
    )  # 하늘색
    fill_category = PatternFill(
        start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"
    )  # 연회색
    fill_subtotal = PatternFill(
        start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"
    )  # 노란색
    fill_total = PatternFill(
        start_color="FDE68A", end_color="FDE68A", fill_type="solid"
    )  # 진한 노란색
    fill_note_head = PatternFill(
        start_color="F8FAF6", end_color="F8FAF6", fill_type="solid"
    )  # 연두색

    # ========================================================================
    # 컬럼 너비 설정 (9개 컬럼)
    # ========================================================================
    widths = [6, 18, 25, 20, 8, 8, 14, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ========================================================================
    # Row 1: 타이틀
    # ========================================================================
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "상 세 견 적 서 (을 지)"
    title_cell.font = font_title
    title_cell.alignment = align_center

    # ========================================================================
    # Row 2: 테이블 헤더
    # ========================================================================
    headers = [
        "No",
        "장비명",
        "품명",
        "규격",
        "수량",
        "단위",
        "단가",
        "공급가액",
        "비고",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = border_base

    # ========================================================================
    # 데이터 및 소계 렌더링
    # ========================================================================
    resources = data.get("resources", [])
    groups = group_by_major_then_machine(resources)

    # major 순서 정의 (JavaScript와 동일)
    major_order = ["자재비", "인건비", "출장경비", "관리비"]

    curr_row = 3
    global_no = 1
    total_sum = 0

    def render_section(major_key: str):
        """각 major 섹션을 렌더링하는 함수"""
        nonlocal curr_row, global_no, total_sum

        if major_key not in groups:
            return

        # 섹션 타이틀 (카테고리 헤더)
        display_title = MAJOR_DISPLAY_MAP.get(major_key, f"{major_key} 상세 내역")
        ws.merge_cells(f"A{curr_row}:I{curr_row}")
        category_cell = ws.cell(curr_row, 1)
        category_cell.value = f"■ {display_title}"
        category_cell.font = font_bold
        category_cell.fill = fill_category
        category_cell.border = border_base
        category_cell.alignment = align_left
        curr_row += 1

        # 각 machine 그룹 렌더링
        machines = groups[major_key]
        major_total = 0

        for machine_name, items in machines.items():
            start_machine_row = curr_row

            for item in items:
                compare = item.get("compare", 0)
                solo_price = item.get("solo_price", 0)
                subtotal = compare * solo_price
                major_total += subtotal
                total_sum += subtotal

                # No
                cell = ws.cell(curr_row, 1, global_no)
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base
                global_no += 1

                # 장비명 (첫 행에만, 나중에 병합)
                cell = ws.cell(curr_row, 2, machine_name)
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 품명
                cell = ws.cell(curr_row, 3, item.get("minor", ""))
                cell.font = font_default
                cell.alignment = align_left
                cell.border = border_base

                # 규격
                cell = ws.cell(curr_row, 4, item.get("spec", ""))
                cell.font = font_default
                cell.alignment = align_left
                cell.border = border_base

                # 수량
                cell = ws.cell(curr_row, 5, compare)
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 단위
                cell = ws.cell(curr_row, 6, item.get("unit", "식"))
                cell.font = font_default
                cell.alignment = align_center
                cell.border = border_base

                # 단가
                cell = ws.cell(curr_row, 7, solo_price)
                cell.font = font_default
                cell.alignment = align_right
                cell.border = border_base
                cell.number_format = "#,##0"

                # 공급가액
                cell = ws.cell(curr_row, 8, subtotal)
                cell.font = font_bold
                cell.alignment = align_right
                cell.border = border_base
                cell.number_format = "#,##0"

                # 비고
                cell = ws.cell(curr_row, 9, item.get("description", ""))
                cell.font = font_default
                cell.alignment = align_left
                cell.border = border_base

                curr_row += 1

            # 장비명 셀 병합 (동일 장비의 여러 행)
            if len(items) > 1:
                ws.merge_cells(f"B{start_machine_row}:B{curr_row - 1}")

        # major 소계 행
        ws.merge_cells(f"A{curr_row}:G{curr_row}")
        subtotal_cell = ws.cell(curr_row, 1)
        subtotal_cell.value = f"{major_key} 총 합계"
        subtotal_cell.font = font_bold
        subtotal_cell.alignment = align_center
        subtotal_cell.fill = fill_subtotal
        subtotal_cell.border = border_subtotal

        # 소계 금액
        for col in range(2, 8):
            cell = ws.cell(curr_row, col)
            cell.fill = fill_subtotal
            cell.border = border_subtotal

        amount_cell = ws.cell(curr_row, 8, major_total)
        amount_cell.font = font_bold
        amount_cell.alignment = align_right
        amount_cell.fill = fill_subtotal
        amount_cell.border = border_subtotal
        amount_cell.number_format = "#,##0"

        # 비고 셀
        ws.cell(curr_row, 9).fill = fill_subtotal
        ws.cell(curr_row, 9).border = border_subtotal

        curr_row += 1

    # 정해진 순서대로 렌더링
    for major in major_order:
        render_section(major)

    # majorOrder에 없는 기타 카테고리도 렌더링
    for major in groups.keys():
        if major not in major_order:
            render_section(major)

    # ========================================================================
    # 최종 합계
    # ========================================================================
    ws.merge_cells(f"A{curr_row}:G{curr_row}")
    total_cell = ws.cell(curr_row, 1)
    total_cell.value = "합 계 (VAT 별도)"
    total_cell.font = font_bold
    total_cell.alignment = align_center
    total_cell.fill = fill_total
    total_cell.border = border_base

    for col in range(2, 8):
        cell = ws.cell(curr_row, col)
        cell.fill = fill_total
        cell.border = border_base

    total_amount_cell = ws.cell(curr_row, 8, total_sum)
    total_amount_cell.font = font_bold
    total_amount_cell.alignment = align_right
    total_amount_cell.fill = fill_total
    total_amount_cell.border = border_base
    total_amount_cell.number_format = "#,##0"

    ws.cell(curr_row, 9).fill = fill_total
    ws.cell(curr_row, 9).border = border_base

    curr_row += 1

    # ========================================================================
    # 빈 행
    # ========================================================================
    curr_row += 1

    # ========================================================================
    # 비고 섹션
    # ========================================================================
    ws.merge_cells(f"A{curr_row}:I{curr_row}")
    note_head_cell = ws.cell(curr_row, 1)
    note_head_cell.value = "비고 (Note)"
    note_head_cell.font = font_bold
    note_head_cell.alignment = align_center
    note_head_cell.fill = fill_note_head
    note_head_cell.border = border_base
    curr_row += 1

    # 비고 내용 (4행 병합)
    ws.merge_cells(f"A{curr_row}:I{curr_row + 3}")
    note_content = data.get("description", "특이사항 없음")
    note_cell = ws.cell(curr_row, 1)
    note_cell.value = note_content
    note_cell.font = font_default
    note_cell.alignment = align_top_wrap
    note_cell.border = border_base

    # 병합된 셀의 테두리 적용
    for row in range(curr_row, curr_row + 4):
        for col in range(1, 10):
            ws.cell(row, col).border = border_base

    # ========================================================================
    # Excel 파일 생성 및 반환
    # ========================================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
