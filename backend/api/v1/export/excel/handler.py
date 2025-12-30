# backend/api/v1/export/excel/handler.py
import asyncio
from datetime import datetime
from io import BytesIO
from pathlib import Path
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session

from backend.database import get_db

# PDF 모듈의 로직 재활용
from ..pdf.crud import (
    get_save_file_path,
    load_settings,
    open_file_in_explorer,
    sanitize_filename,
)
from . import crud
from .Format import detailed, header, price_compare

handler = APIRouter()


# ============================================================================
# 폴더 통합 Excel Export (먼저 정의해야 함 - 라우트 우선순위)
# ============================================================================

@handler.get("/folder/{folder_id}")
async def export_folder_excel(folder_id: UUID, db: Session = Depends(get_db)):
    """
    폴더의 모든 견적서를 하나의 Excel 파일로 통합 다운로드
    순서: 갑지(Header), 을지(Detailed), 내정가비교서(PriceCompare)
    각 문서는 별도의 시트로 생성됨
    """
    try:
        # 1. 폴더 정보 가져오기
        from backend.api.v1.quotation.folder.crud import get_folder_by_id

        folder = get_folder_by_id(db, folder_id)
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")

        # 2. 폴더의 각 리소스 조회
        from backend.api.v1.quotation.folder.crud import get_folder_resources

        resources = get_folder_resources(db, folder_id)

        header_resource = next(
            (r for r in resources if r["table_name"] == "견적서"), None
        )
        detailed_resource = next(
            (r for r in resources if r["table_name"] == "견적서(을지)"), None
        )
        price_compare_resource = next(
            (r for r in resources if r["table_name"] == "내정가 비교"), None
        )

        # 리소스가 하나도 없으면 에러 반환
        if not header_resource and not detailed_resource and not price_compare_resource:
            raise HTTPException(
                status_code=400,
                detail="폴더에 생성된 견적서가 없습니다. 갑지, 을지, 또는 내정가비교서를 먼저 생성해주세요."
            )

        # 3. 각 문서의 Excel 파일을 개별적으로 생성하여 시트로 병합
        import openpyxl
        from copy import copy

        print(f"[폴더 Excel Export] 갑지: {header_resource is not None}")
        print(f"[폴더 Excel Export] 을지: {detailed_resource is not None}")
        print(f"[폴더 Excel Export] 내정가비교서: {price_compare_resource is not None}")

        def copy_sheet_to_workbook(source_wb, target_wb, sheet_title):
            """다른 워크북에서 시트를 복사하는 헬퍼 함수"""
            source_sheet = source_wb.active
            target_sheet = target_wb.create_sheet(title=sheet_title)

            # 셀 값과 스타일 복사
            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet[cell.coordinate]
                    target_cell.value = cell.value

                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = copy(cell.number_format)
                        target_cell.protection = copy(cell.protection)
                        target_cell.alignment = copy(cell.alignment)

            # 병합된 셀 복사
            for merged_cell_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_cell_range))

            # 열 너비 복사
            for col_letter, col_dim in source_sheet.column_dimensions.items():
                target_sheet.column_dimensions[col_letter].width = col_dim.width

            # 행 높이 복사
            for row_num, row_dim in source_sheet.row_dimensions.items():
                target_sheet.row_dimensions[row_num].height = row_dim.height

            # 페이지 설정 복사
            target_sheet.page_setup = copy(source_sheet.page_setup)
            target_sheet.print_options = copy(source_sheet.print_options)

            return target_sheet

        output = BytesIO()
        wb = None

        # 4. 각 문서를 시트로 추가 (순서: 갑지, 을지, 내정가비교서)

        # 4-1. 갑지 시트
        if header_resource:
            try:
                print(f"[폴더 Excel Export] 갑지 생성 시작: {header_resource['id']}")
                resource_id = header_resource["id"]
                if isinstance(resource_id, str):
                    resource_id = UUID(resource_id)

                header_data = crud.get_header_data(db, resource_id)
                header_wb = header.create_excel(header_data)
                header_wb.seek(0)
                temp_wb = openpyxl.load_workbook(header_wb)

                if wb is None:
                    wb = temp_wb
                    wb.active.title = "갑지"
                else:
                    copy_sheet_to_workbook(temp_wb, wb, "갑지")
                print(f"[폴더 Excel Export] 갑지 생성 완료")
            except Exception as e:
                print(f"[폴더 Excel Export] 갑지 시트 생성 오류: {e}")
                import traceback
                traceback.print_exc()

        # 4-2. 을지 시트
        if detailed_resource:
            try:
                print(f"[폴더 Excel Export] 을지 생성 시작: {detailed_resource['id']}")
                resource_id = detailed_resource["id"]
                if isinstance(resource_id, str):
                    resource_id = UUID(resource_id)

                detailed_data = crud.get_detailed_data(db, resource_id)
                detailed_wb = detailed.create_excel(detailed_data)
                detailed_wb.seek(0)
                temp_wb = openpyxl.load_workbook(detailed_wb)

                if wb is None:
                    wb = temp_wb
                    wb.active.title = "을지"
                else:
                    copy_sheet_to_workbook(temp_wb, wb, "을지")
                print(f"[폴더 Excel Export] 을지 생성 완료")
            except Exception as e:
                print(f"[폴더 Excel Export] 을지 시트 생성 오류: {e}")
                import traceback
                traceback.print_exc()

        # 4-3. 내정가비교서 시트
        if price_compare_resource:
            try:
                print(f"[폴더 Excel Export] 내정가비교서 생성 시작: {price_compare_resource['id']}")
                resource_id = price_compare_resource["id"]
                if isinstance(resource_id, str):
                    resource_id = UUID(resource_id)

                pc_data = crud.get_price_compare_data(db, resource_id)
                pc_wb = price_compare.create_excel(pc_data)
                pc_wb.seek(0)
                temp_wb = openpyxl.load_workbook(pc_wb)

                if wb is None:
                    wb = temp_wb
                    wb.active.title = "내정가비교서"
                else:
                    copy_sheet_to_workbook(temp_wb, wb, "내정가비교서")
                print(f"[폴더 Excel Export] 내정가비교서 생성 완료")
            except Exception as e:
                print(f"[폴더 Excel Export] 내정가비교서 시트 생성 오류: {e}")
                import traceback
                traceback.print_exc()

        # 5. 최종 워크북 저장
        if wb is None:
            raise HTTPException(
                status_code=500, detail="워크북 생성 실패"
            )

        wb.save(output)
        output.seek(0)

        # 6. 파일명 생성 및 파일 시스템에 저장
        import asyncio

        settings = load_settings()
        base_path = settings.get("pdfSavePath") or str(
            Path.home() / "Documents" / "JLT_견적서"
        )
        ask_location = settings.get("askSaveLocation", False)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        raw_filename = f"{folder.title}_통합견적서_{timestamp}.xlsx"
        safe_filename = sanitize_filename(raw_filename)

        if ask_location:
            # 수동 저장: 파일 대화상자 열기
            loop = asyncio.get_event_loop()
            file_path = await loop.run_in_executor(
                None, lambda: get_save_file_path(safe_filename, base_path)
            )

            if not file_path:
                return JSONResponse(
                    {"success": False, "message": "저장이 취소되었습니다."}
                )

            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"
        else:
            # 자동 저장: /{general.name}/{folder.title}/
            safe_general_name = sanitize_filename(folder.general.name, "견적서")
            safe_folder_title = sanitize_filename(folder.title, "폴더")
            save_dir = Path(base_path) / safe_general_name / safe_folder_title
            save_dir.mkdir(parents=True, exist_ok=True)
            file_path = str(save_dir / safe_filename)

        # 7. 파일 물리적 저장
        with open(file_path, "wb") as f:
            f.write(output.getvalue())

        # 8. 탐색기 열기 (자동 저장일 때만)
        if not ask_location:
            open_file_in_explorer(file_path)

        return JSONResponse({"success": True, "path": file_path})

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Folder Excel export error: {str(e)}"
        )


# ============================================================================
# 개별 견적서 Excel Export
# ============================================================================

@handler.get("/{quotation_type}/{quotation_id}")
async def export_excel(
    quotation_type: str, quotation_id: UUID, db: Session = Depends(get_db)
):
    """
    견적서 Excel 내보내기 (PDF 저장 설정 및 로직 적용)
    """
    # 1. quotation_type 검증
    valid_types = ["header", "detailed", "price_compare"]
    if quotation_type not in valid_types:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid quotation_type. Must be one of {valid_types}",
        )

    # 2. 데이터 조회
    try:
        if quotation_type == "header":
            data = crud.get_header_data(db, quotation_id)
            type_name = "갑지"
        elif quotation_type == "detailed":
            data = crud.get_detailed_data(db, quotation_id)
            type_name = "을지"
        elif quotation_type == "price_compare":
            data = crud.get_price_compare_data(db, quotation_id)
            type_name = "내역"
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))

    # 3. Excel 파일 생성 (바이너리)
    try:
        if quotation_type == "header":
            excel_file = header.create_excel(data)
        elif quotation_type == "detailed":
            excel_file = detailed.create_excel(data)
        elif quotation_type == "price_compare":
            excel_file = price_compare.create_excel(data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation error: {str(e)}")

    # 4. 저장 경로 및 파일명 결정 (견적서(일반)/폴더명/Excel 구조 적용)
    try:
        settings = load_settings()
        base_path = settings.get("pdfSavePath") or str(
            Path.home() / "Documents" / "JLT_견적서"
        )
        ask_location = settings.get("askSaveLocation", False)

        # 파일명 생성 및 안전한 이름 변환
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        raw_filename = f"견적서({type_name})_{timestamp}.xlsx"
        safe_filename = sanitize_filename(raw_filename)

        # 💡 폴더 정보를 통해 견적서(일반)명과 폴더명 가져오기
        from backend.api.v1.quotation.folder.crud import get_folder_by_id
        from backend.models.general import General

        folder_id = None
        general_name = None
        folder_title = None

        # 각 견적서 타입에서 folder_id 추출
        if quotation_type == "header" and data:
            folder_id = data.get("folder_id")
        elif quotation_type == "detailed" and data:
            folder_id = data.get("folder_id")
        elif quotation_type == "price_compare" and data:
            folder_id = data.get("folder_id")

        # folder_id로 폴더 및 견적서(일반) 정보 가져오기
        if folder_id:
            folder = get_folder_by_id(db, UUID(str(folder_id)))
            if folder:
                folder_title = folder.title
                general = db.query(General).filter(General.id == folder.general_id).first()
                if general:
                    general_name = general.name

        if ask_location:
            # 💡 Windows 대화상자 사용 (Executor로 비동기 처리)
            loop = asyncio.get_event_loop()
            file_path = await loop.run_in_executor(
                None,
                lambda: get_save_file_path(
                    safe_filename, base_path
                ),  # PDF crud의 함수 재활용
            )

            if not file_path:
                return JSONResponse(
                    {"success": False, "message": "저장이 취소되었습니다."}
                )

            # 확장자가 없으면 자동으로 붙여줌
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"
        else:
            # 💡 개별 저장: /{general.name}/{folder.title}/Excel/
            if general_name and folder_title:
                safe_general_name = sanitize_filename(general_name, "견적서")
                safe_folder_title = sanitize_filename(folder_title, "폴더")
                save_dir = Path(base_path) / safe_general_name / safe_folder_title / "Excel"
            elif general_name:
                # Fallback: 폴더명 없는 경우
                safe_general_name = sanitize_filename(general_name, "견적서")
                save_dir = Path(base_path) / safe_general_name / "Excel"
            else:
                # Fallback: general.name도 없는 경우
                save_dir = Path(base_path) / "Excel"

            # 폴더가 없으면 생성 (이미 존재하면 skip)
            save_dir.mkdir(parents=True, exist_ok=True)
            file_path = str(save_dir / safe_filename)

        # 5. 파일 물리적 저장
        with open(file_path, "wb") as f:
            f.write(excel_file.getvalue())

        # 6. 탐색기 열기 (위치 질문이 아닐 때만)
        if not ask_location:
            open_file_in_explorer(file_path)

        # 7. 응답 반환 (JSON)
        return JSONResponse(
            {
                "success": True,
                "path": file_path,
                "message": "성공적으로 저장되었습니다.",
            }
        )

    except Exception as e:
        return JSONResponse({"success": False, "message": f"File save error: {str(e)}"})
