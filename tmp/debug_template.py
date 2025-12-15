import os
from openpyxl import load_workbook

EXCEL_FILE_PATH = "tmp/SYNEX+QUOTATION INFO (1).xlsx"
SHEET_NAME_HINT = "1st EL_1st Fill" 

def debug_sheet_raw():
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"File not found: {EXCEL_FILE_PATH}")
        return

    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
    target_ws = None
    
    for name in wb.sheetnames:
        if SHEET_NAME_HINT in name:
            target_ws = wb[name]
            break
            
    if not target_ws:
        print(f"Sheet not found. Available: {wb.sheetnames}")
        return

    print(f"\n=== Sheet: {target_ws.title} Raw Dump ===")
    
    # 처음 5줄만 무조건 찍어봄 (헤더 위치 확인용)
    for i, row in enumerate(target_ws.iter_rows(max_row=5, values_only=True), 1):
        print(f"Row {i}: {row}")

    print("\n--- Labor Section Check (Rows 130~150) ---")
    # 인건비가 있을법한 위치 집중 확인
    for i, row in enumerate(target_ws.iter_rows(min_row=130, max_row=150, values_only=True), 130):
        # 값이 있는 행만 출력
        if any(row):
            # 보기 편하게 인덱스와 함께 출력
            row_display = []
            for col_idx, cell in enumerate(row):
                if cell: row_display.append(f"[{col_idx}]{cell}")
            print(f"Row {i}: {', '.join(row_display)}")

if __name__ == "__main__":
    debug_sheet_raw()