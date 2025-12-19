# register_parts.py (openpyxl ê¸°ë°˜)
import json
import os
from openpyxl import load_workbook
import urllib.request
import urllib.error

# --- Configuration ---
API_URL = "http://localhost:8000/api/v1/parts"
EXCEL_FILE_PATH = "tmp/data.xlsx"
# í˜„ì¬ ì—‘ì…€: 1ë²ˆ ì‹œíŠ¸(0-based 0)ê°€ parts
PARTS_SHEET_NAME = 0 

def clean_value(value):
    """Excel í•„ë“œ ê°’ì„ ì •ë¦¬í•˜ê³  'ê³µë°±', 'nan' ë“±ì„ ë¹ˆ ë¬¸ìì—´("")ë¡œ ë³€í™˜í•©ë‹ˆë‹¤."""
    if value is None:
        return ""

    cleaned = str(value).strip()
    
    # âœ… ìˆ˜ì •: 'ê³µë°±' í…ìŠ¤íŠ¸ì´ê±°ë‚˜ 'nan', 'None', ë˜ëŠ” ì™„ì „íˆ ë¹ˆ ë¬¸ìì—´ì¸ ê²½ìš° ëª¨ë‘ ë¹ˆ ë¬¸ìì—´("") ë°˜í™˜
    if cleaned in ("ê³µë°±", "nan", "None") or not cleaned:
        return ""
        
    return cleaned

def clean_maker_name(value):
    """Maker nameì€ 'ê³µë°±'ì„ ' 'ë¡œ ìœ ì§€(ì œì¡°ì‚¬ëª… ë§¤ì¹­ìš©)."""
    if value is None:
        return ""

    cleaned = str(value).strip()
    if cleaned == "ê³µë°±":
        return " "
    if cleaned in ("nan", "None") or not cleaned:
        return ""
    return cleaned

def get_parts_data_from_excel(excel_file_path, sheet_name):
    """Excel íŒŒì¼ì„ ì½ì–´ Parts ë“±ë¡ì— í•„ìš”í•œ JSON í˜•ì‹ì˜ ë¦¬ìŠ¤íŠ¸ë¡œ ë³€í™˜í•©ë‹ˆë‹¤."""
    data_list = []
    skipped_count = 0 
    
    if not os.path.exists(excel_file_path):
        print(f"ì˜¤ë¥˜: íŒŒì¼ ê²½ë¡œë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {excel_file_path}")
        return [], 0

    try:
        wb = load_workbook(excel_file_path, data_only=True)
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
    except Exception as e:
        print(f"ì˜¤ë¥˜: ì—‘ì…€ íŒŒì¼/ì‹œíŠ¸({sheet_name}) ì½ê¸° ì‹¤íŒ¨: {e}")
        return [], 0

    # í˜„ì¬ parts ì‹œíŠ¸ í—¤ë”
    # Unit | í’ˆëª© | ëª¨ë¸ëª…/ê·œê²© | ì œì¡°ì‚¬ | UL | CE | KC | ê¸°íƒ€ | ë‹¨ìœ„ | ë‹¨ê°€
    required_cols = ['Unit', 'í’ˆëª©', 'ëª¨ë¸ëª…/ê·œê²©', 'ì œì¡°ì‚¬', 'UL', 'CE', 'KC', 'ê¸°íƒ€', 'ë‹¨ìœ„', 'ë‹¨ê°€']

    header_row_index = 1  # pandas header=0 ê³¼ ë™ì¼(1ë²ˆì§¸ ì¤„ì´ í—¤ë”)
    header_cells = [cell.value for cell in ws[header_row_index]]
    header = [str(v).strip() if v is not None else "" for v in header_cells]
    header_map = {name: idx for idx, name in enumerate(header)}

    if not all(col in header_map for col in required_cols):
        return [], 0

    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        # clean_valueë¥¼ ì ìš©í•˜ì§€ ì•Šì€ ì›ë³¸ ëª¨ë¸ëª…/ê·œê²© ê°’ (ë¹„ì–´ìˆëŠ”ì§€, '-'ì¸ì§€ í™•ì¸ìš©)
        raw_part_name = str(row[header_map['ëª¨ë¸ëª…/ê·œê²©']] if header_map['ëª¨ë¸ëª…/ê·œê²©'] < len(row) else "").strip()
        
        # âš ï¸ í•„í„°ë§ ì¡°ê±´: ì›ë³¸ ëª¨ë¸ëª…/ê·œê²©ì´ '-'ì¸ ê²½ìš°ë§Œ ê±´ë„ˆëœë‹ˆë‹¤.
        if raw_part_name == '-': 
            skipped_count += 1
            continue

        try:
            # ëª¨ë“  í•„ë“œì— clean_value ì ìš© (ë¹ˆ ê°’ì€ ""ê°€ ë¨)
            part_name = clean_value(row[header_map['ëª¨ë¸ëª…/ê·œê²©']] if header_map['ëª¨ë¸ëª…/ê·œê²©'] < len(row) else None)
            maker_name = clean_maker_name(row[header_map['ì œì¡°ì‚¬']] if header_map['ì œì¡°ì‚¬'] < len(row) else None)
            major_category = clean_value(row[header_map['Unit']] if header_map['Unit'] < len(row) else None)
            minor_category = clean_value(row[header_map['í’ˆëª©']] if header_map['í’ˆëª©'] < len(row) else None)
            
            ul_status = clean_value(row[header_map['UL']] if header_map['UL'] < len(row) else None) == 'O'
            ce_status = clean_value(row[header_map['CE']] if header_map['CE'] < len(row) else None) == 'O'
            kc_status = clean_value(row[header_map['KC']] if header_map['KC'] < len(row) else None) == 'O'
            
            price_str = clean_value(row[header_map['ë‹¨ê°€']] if header_map['ë‹¨ê°€'] < len(row) else None).replace(',', '').strip()
            # price_strì´ "" (ë¹ˆ ë¬¸ìì—´)ì´ ë˜ëŠ” ê²½ìš°ë¥¼ ëŒ€ë¹„í•˜ì—¬ isdigit() ê²€ì‚¬ ì „ì— strip()
            solo_price = int(price_str.strip()) if price_str.strip().isdigit() else 0
            
            part_json = {
                "maker_name": maker_name,
                "major_category": major_category,
                "minor_category": minor_category,
                "name": part_name,
                "unit": clean_value(row[header_map['ë‹¨ìœ„']] if header_map['ë‹¨ìœ„'] < len(row) else "ea") or "ea",
                "solo_price": solo_price,
                "display_order": len(data_list),
                "ul": ul_status,
                "ce": ce_status,
                "kc": kc_status,
                "certification_etc": clean_value(row[header_map['ê¸°íƒ€']] if header_map['ê¸°íƒ€'] < len(row) else None),
            }
            data_list.append(part_json)

        except Exception as e:
            skipped_count += 1
            pass

    return data_list, skipped_count

def post_json(url: str, payload: dict) -> tuple[int, str]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            return resp.getcode(), body
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return e.code, body

def post_parts_to_api(data_list, api_url):
    """ë³€í™˜ëœ ë¶€í’ˆ ë°ì´í„° ë¦¬ìŠ¤íŠ¸ë¥¼ APIì— POST ìš”ì²­ìœ¼ë¡œ ì „ì†¡í•©ë‹ˆë‹¤. (ì´ì „ê³¼ ë™ì¼)"""
    print(f"\n--- ë¶€í’ˆ API ì „ì†¡ ì‹œì‘ ({len(data_list)}ê°œ ë¶€í’ˆ) ---")
    
    success_count = 0
    failure_count = 0
    
    for part_data in data_list:
        try:
            status_code, body = post_json(api_url, part_data)

            if 200 <= status_code < 300:
                success_count += 1
            else:
                failure_count += 1
                response_text = body
                if len(response_text) > 100:
                    response_text = response_text[:100] + "..."
                # ë¶€í’ˆëª…ì— ê³µë°±ì´ ìˆì„ ìˆ˜ ìˆìœ¼ë¯€ë¡œ repr()ë¡œ ì¶œë ¥í•˜ì—¬ ëª…í™•í•˜ê²Œ í•©ë‹ˆë‹¤.
                print(f"âŒ ì‹¤íŒ¨: {repr(part_data['name'])} (Status: {status_code}, ì‘ë‹µ: {response_text})")
                
        except urllib.error.URLError as e:
            failure_count += 1
            print(f"ğŸš¨ ì—°ê²° ì˜¤ë¥˜: {repr(part_data['name'])} ì „ì†¡ ì‹¤íŒ¨. ({e})")

    print(f"\n--- API ì „ì†¡ ì™„ë£Œ ---")
    print(f"âœ… ì„±ê³µì ìœ¼ë¡œ ë“±ë¡ëœ ë¶€í’ˆ: {success_count}ê°œ")
    print(f"âŒ ë“±ë¡ ì‹¤íŒ¨/ì˜¤ë¥˜ ë¶€í’ˆ: {failure_count}ê°œ")


if __name__ == "__main__":
    # âš ï¸ ì¤‘ìš”: register_makers.pyë¥¼ ë¨¼ì € ì‹¤í–‰í•˜ì—¬ ë¹ˆ ì œì¡°ì‚¬("")ê°€ ë“±ë¡ë˜ì—ˆëŠ”ì§€ í™•ì¸í•˜ì„¸ìš”.
    all_parts_data, skipped_count = get_parts_data_from_excel(EXCEL_FILE_PATH, PARTS_SHEET_NAME)

    if all_parts_data:
        total_converted = len(all_parts_data)
        print(f"ì´ {total_converted}ê°œì˜ ë¶€í’ˆ ë°ì´í„°ê°€ JSON í˜•ì‹ìœ¼ë¡œ ë³€í™˜ë˜ì—ˆìŠµë‹ˆë‹¤.")
        # ì œì™¸ëœ í–‰ì€ '-'ì¸ ê²½ìš°ë§Œ ë‚¨ìŠµë‹ˆë‹¤.
        print(f"â„¹ï¸ {skipped_count}ê°œì˜ í–‰ì´ 'ë¶€í’ˆëª…'ì´ '-'ì—¬ì„œ ì œì™¸ë˜ì—ˆìŠµë‹ˆë‹¤. (ì´ ë³€í™˜ ì‹œë„: {total_converted + skipped_count}ê°œ)")
        post_parts_to_api(all_parts_data, API_URL)
    else:
        print("ğŸš¨ ë³€í™˜ëœ ë°ì´í„°ê°€ ì—†ì–´ API ì „ì†¡ì„ ì‹œì‘í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤.")
