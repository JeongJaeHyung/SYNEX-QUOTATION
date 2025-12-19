# register_templates.py (openpyxl 기반)
#
# 목적:
# - 하나의 엑셀 파일에서 여러 "템플릿" 시트를 읽어 Machine(견적)과 MachineResources를 생성합니다.
# - 시트의 "수량"과 "단가"를 그대로 MachineResources(quantity, solo_price)로 저장합니다.
# - 부품 식별은 DB에 이미 등록된 Parts(Resources)를 기준으로 (제조사/Unit/품목/모델명) 매칭합니다.
#
# 사용 전제:
# 1) 제조사/부품(마스터)이 먼저 등록되어 있어야 합니다. (tmp/register_makers.py → tmp/register_parts.py)
# 2) FastAPI 서버가 실행 중이어야 합니다. (기본: http://localhost:8005)
#
# 주의:
# - 현재 백엔드에 "is_template" 같은 플래그가 없어서, 일반 Machine으로 등록됩니다.
#   구분이 필요하면 MACHINE_NAME_PREFIX / TEMPLATE_CREATOR 값을 바꿔서 식별하세요.
import json
import os
import urllib.parse
import urllib.request
import urllib.error
from openpyxl import load_workbook
from datetime import datetime

# --- Configuration ---
API_BASE_URL = "http://localhost:8000/api/v1"
MACHINE_API_URL = f"{API_BASE_URL}/quotation/machine/"
MACHINE_SEARCH_URL = f"{API_BASE_URL}/quotation/machine/search"
PARTS_LIST_URL = f"{API_BASE_URL}/parts"
PARTS_CREATE_URL = f"{API_BASE_URL}/parts"
MAKER_CREATE_URL = f"{API_BASE_URL}/maker"

EXCEL_FILE_PATH = "tmp/data.xlsx"

# 엑셀 시트 인덱스(0-based): 1~3번 시트는 마스터, 4~9번 시트는 템플릿(= 3~8)
TEMPLATE_SHEET_INDICES = list(range(3, 9))

MACHINE_NAME_PREFIX = "[TEMPLATE] "
TEMPLATE_CREATOR = "TEMPLATE"

# 템플릿 등록 전에 마스터(Parts)에 없는 부품은 템플릿 행 기반으로 자동 생성할지 여부
AUTO_CREATE_MISSING_PARTS = True

# 템플릿 name이 이미 존재하면 새로 만들지 않고 PUT으로 갱신할지 여부
UPSERT_TEMPLATES = True

# 기존에 같은 이름 템플릿이 여러 개 있으면(과거 중복 생성), 오래된 것들을 자동으로 이름 변경하여 중복을 해소
DEDUPLICATE_EXISTING_TEMPLATES = True

# 매칭 누락이 있어도 가능한 것만 등록할지 여부
# - True: 매칭 실패가 1건이라도 있으면 해당 시트 등록 안 함
# - False: 매칭 성공한 항목만으로 템플릿을 등록(테스트/점진 보완용)
STRICT_MODE = False

SUMMARY_MAJORS = {"전장/제어부 집계"}
LABOR_MAJORS = {"인건비"}


def clean_value(value):
    if value is None:
        return ""
    cleaned = str(value).strip()
    if cleaned in ("공백", "nan", "None") or not cleaned:
        return ""
    return cleaned


def clean_maker_name(value):
    if value is None:
        return ""
    cleaned = str(value).strip()
    if cleaned == "공백":
        return " "
    if cleaned in ("nan", "None") or not cleaned:
        return ""
    return cleaned


def normalize_key(value: str) -> str:
    """매칭용 정규화: 공백 정리 + case-insensitive."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return " ".join(text.split()).casefold()


def normalize_loose(value: str) -> str:
    """더 느슨한 매칭용: 영숫자만 남김."""
    base = normalize_key(value)
    if not base:
        return ""
    return "".join(ch for ch in base if ch.isalnum())


def parse_int(value, default=0) -> int:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except Exception:
            return default
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        return int(float(text))
    except Exception:
        return default


def get_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        body = resp.read().decode("utf-8", errors="replace")
        return json.loads(body)


def post_json(url: str, payload: dict) -> tuple[int, str]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            return resp.getcode(), body
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return e.code, body


def put_json(url: str, payload: dict) -> tuple[int, str]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="PUT",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            return resp.getcode(), body
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return e.code, body


def create_maker_if_missing(maker_name: str) -> bool:
    maker_name = clean_maker_name(maker_name)
    if not maker_name:
        return False
    # " " (공백 제조사 / T000) 는 자동 생성하면 ID가 달라질 수 있으니 시도하지 않음
    if maker_name == " ":
        return False
    status, body = post_json(MAKER_CREATE_URL, {"name": maker_name})
    if 200 <= status < 300:
        return True
    # 이미 존재하는 경우(409)도 OK 처리
    if status == 409:
        return True
    return False


def parse_cert_bool(value) -> bool:
    return clean_value(value) == "O"


def add_part_to_index(parts_index: dict, part: dict) -> None:
    maker_name = clean_maker_name(part.get("maker_name"))
    major = clean_value(part.get("major_category"))
    minor = clean_value(part.get("minor_category"))
    name = clean_value(part.get("name"))
    part_payload = {
        "maker_id": part.get("maker_id"),
        "resources_id": part.get("id"),
        "solo_price": part.get("solo_price", 0) or 0,
        "unit": part.get("unit") or "",
    }

    maker_k = normalize_key(maker_name)
    name_k = normalize_key(name)
    if maker_k and name_k:
        parts_index["by_maker_model"].setdefault((maker_k, name_k), part_payload)

        loose_key = (normalize_loose(maker_name), normalize_loose(name))
        if all(loose_key):
            parts_index["by_maker_model_loose"].setdefault(loose_key, part_payload)

    if not name_k and major and minor:
        key_mm = (normalize_key(major), normalize_key(minor), normalize_key(name))
        if key_mm[0] and key_mm[1]:
            parts_index["by_major_minor_empty_name"].setdefault(key_mm, part_payload)


def create_part_from_template_row(parts_index: dict, row_payload: dict) -> dict | None:
    """
    템플릿 row를 기반으로 Parts를 생성하고, 성공 시 parts_index에 반영합니다.
    row_payload 키: maker_name, major_category, minor_category, name, unit, solo_price, ul, ce, kc, certification_etc
    """
    if not AUTO_CREATE_MISSING_PARTS:
        return None

    maker_name = clean_maker_name(row_payload.get("maker_name"))
    major = clean_value(row_payload.get("major_category"))
    minor = clean_value(row_payload.get("minor_category"))
    name = clean_value(row_payload.get("name"))

    # summary/labor는 maker가 비어있으면 " "로 강제 (T000 maker_name)
    if major in (SUMMARY_MAJORS | LABOR_MAJORS) and not maker_name:
        maker_name = " "

    # 일반 부품은 maker/name이 없으면 생성 불가
    if major not in (SUMMARY_MAJORS | LABOR_MAJORS):
        if not maker_name or not name:
            return None

    # [수정] 인건비/집계 항목인데 이름이 없으면 품목(minor)명을 이름으로 사용
    if major in (SUMMARY_MAJORS | LABOR_MAJORS) and not name:
        name = minor

    payload = {
        "maker_name": maker_name,
        "major_category": major,
        "minor_category": minor,
        "name": name,  # labor/summary는 "" 가능
        "unit": clean_value(row_payload.get("unit")) or ("M/D" if major in LABOR_MAJORS else "ea"),
        "solo_price": int(row_payload.get("solo_price") or 0),
        # display_order는 비워두면 서버에서 (max+1)로 자동 부여
        "ul": bool(row_payload.get("ul") or False),
        "ce": bool(row_payload.get("ce") or False),
        "kc": bool(row_payload.get("kc") or False),
        "certification_etc": row_payload.get("certification_etc"),
    }

    status, body = post_json(PARTS_CREATE_URL, payload)
    if status == 404 and maker_name and maker_name != " ":
        # maker가 없으면 생성 후 재시도
        if create_maker_if_missing(maker_name):
            status, body = post_json(PARTS_CREATE_URL, payload)

    if not (200 <= status < 300):
        return None

    try:
        part = json.loads(body)
    except Exception:
        return None

    add_part_to_index(parts_index, part)
    return part


def find_existing_machine_id_by_name(machine_name: str) -> str | None:
    # API 제한: limit <= 100
    skip = 0
    limit = 100
    while True:
        query = urllib.parse.urlencode({"search": machine_name, "skip": skip, "limit": limit})
        url = f"{MACHINE_SEARCH_URL}?{query}"
        try:
            data = get_json(url)
        except Exception:
            return None

        items = data.get("items") or []
        for item in items:
            if (item.get("name") or "").strip() == machine_name.strip():
                return item.get("id")

        total = int(data.get("total", 0) or 0)
        skip += limit
        if skip >= total:
            break
    return None


def find_exact_machines_by_name(machine_name: str) -> list[dict]:
    """search API로 후보를 가져온 뒤 name이 정확히 같은 것만 반환합니다."""
    matched: list[dict] = []
    skip = 0
    limit = 100
    while True:
        query = urllib.parse.urlencode({"search": machine_name, "skip": skip, "limit": limit})
        url = f"{MACHINE_SEARCH_URL}?{query}"
        data = get_json(url)
        items = data.get("items") or []
        for item in items:
            if (item.get("name") or "").strip() == machine_name.strip():
                matched.append(item)

        total = int(data.get("total", 0) or 0)
        skip += limit
        if skip >= total:
            break
    return matched


def parse_dt(value: str) -> datetime:
    if not value:
        return datetime.min
    try:
        # 예: 2025-12-14T15:00:00.123456
        return datetime.fromisoformat(value)
    except Exception:
        return datetime.min


def deduplicate_template_name(machine_name: str, template_title: str) -> str | None:
    """
    동일 name이 2개 이상이면:
    - 최신(updated_at max) 1개는 유지
    - 나머지는 이름을 [DUPLICATE] {template_title} ({id}) 로 변경하여 향후 조회/업서트에서 제외되게 함
    반환: 유지할 machine_id (없으면 None)
    """
    if not DEDUPLICATE_EXISTING_TEMPLATES:
        return find_existing_machine_id_by_name(machine_name)

    exact = find_exact_machines_by_name(machine_name)
    if not exact:
        return None
    if len(exact) == 1:
        return exact[0].get("id")

    exact_sorted = sorted(exact, key=lambda x: parse_dt(x.get("updated_at")), reverse=True)
    keep = exact_sorted[0].get("id")
    to_rename = [x.get("id") for x in exact_sorted[1:] if x.get("id")]

    for mid in to_rename:
        new_name = f"[DUPLICATE] {template_title} ({mid})"
        put_json(f"{MACHINE_API_URL}{mid}", {"name": new_name})

    print(f"🧹 템플릿 이름 중복 해소: '{machine_name}' -> keep={keep}, renamed={len(to_rename)}")
    return keep


def fetch_all_parts_index() -> dict:
    """
    Parts(Resources) 전체를 API로 읽어서, 템플릿 시트 row를 DB 부품으로 매칭하기 위한 인덱스를 만듭니다.

    - templates 쪽 Unit/품목 분류가 마스터와 다를 수 있어서,
      1) maker_name + model_name(=name) 기반 인덱스를 우선 사용합니다.
      2) model_name이 비어있는(집계/인건비) 항목은 major + minor + name("")로 매칭합니다.
    """
    by_maker_model: dict[tuple[str, str], dict] = {}
    by_maker_model_loose: dict[tuple[str, str], dict] = {}
    by_major_minor_empty_name: dict[tuple[str, str, str], dict] = {}

    skip = 0
    limit = 1000
    total = None

    while total is None or skip < total:
        query = urllib.parse.urlencode({"skip": skip, "limit": limit})
        url = f"{PARTS_LIST_URL}?{query}"
        data = get_json(url)

        total = int(data.get("total", 0))
        items = data.get("items", []) or []

        for item in items:
            maker_name = clean_maker_name(item.get("maker_name"))
            major = clean_value(item.get("major_category"))
            minor = clean_value(item.get("minor_category"))
            name = clean_value(item.get("name"))

            part_payload = {
                "maker_id": item.get("maker_id"),
                "resources_id": item.get("id"),
                "solo_price": item.get("solo_price", 0) or 0,
                "unit": item.get("unit") or "",
                "name": item.get("name") or "",  # [추가] 이름 정보 저장
            }

            # 1) maker + model (가장 안정적인 매칭)
            maker_k = normalize_key(maker_name)
            name_k = normalize_key(name)
            if maker_k and name_k:
                key = (maker_k, name_k)
                by_maker_model.setdefault(key, part_payload)

                key_loose = (normalize_loose(maker_name), normalize_loose(name))
                if all(key_loose):
                    by_maker_model_loose.setdefault(key_loose, part_payload)

            # 2) 집계/인건비 같은 name="" 항목은 major+minor로 매칭
            if not name_k and major and minor:
                key_mm = (normalize_key(major), normalize_key(minor), normalize_key(name))
                if key_mm[0] and key_mm[1]:
                    by_major_minor_empty_name.setdefault(key_mm, part_payload)

        skip += limit

    return {
        "by_maker_model": by_maker_model,
        "by_maker_model_loose": by_maker_model_loose,
        "by_major_minor_empty_name": by_major_minor_empty_name,
    }


def find_part(parts_index: dict, maker_name: str, major: str, minor: str, name: str) -> dict | None:
    # 1) maker + model (정규화)
    maker_k = normalize_key(maker_name)
    name_k = normalize_key(name)
    if maker_k and name_k:
        found = parts_index["by_maker_model"].get((maker_k, name_k))
        if found:
            return found

        found = parts_index["by_maker_model_loose"].get((normalize_loose(maker_name), normalize_loose(name)))
        if found:
            return found

    # 2) 집계/인건비(name="") 항목: major + minor + name("")로 매칭
    if major and minor and not name_k:
        key_mm = (normalize_key(major), normalize_key(minor), normalize_key(name))
        found = parts_index["by_major_minor_empty_name"].get(key_mm)
        if found:
            return found

    return None


def parse_template_sheet(ws, parts_index: dict) -> tuple[list[dict], list[dict], int]:
    """
    템플릿 시트에서 MachineResources 리스트를 만듭니다.

    시트 헤더(예시):
    Unit | 품목 | 모델명/규격 | 제조사 | UL | CE | KC | 기타 | 수량 | 단위 | 단가
    """
    header_row_index = 1
    header_cells = [cell.value for cell in ws[header_row_index]]
    header = [str(v).strip() if v is not None else "" for v in header_cells]
    header_map = {name: idx for idx, name in enumerate(header)}

    required = ["Unit", "품목", "모델명/규격", "제조사", "수량"]
    if not all(col in header_map for col in required):
        raise ValueError(f"시트 '{ws.title}': 필요한 헤더가 없습니다. 필요={required}, 실제={header}")

    resources: list[dict] = []
    missing: list[dict] = []
    created_parts_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
        maker_name = clean_maker_name(row[header_map["제조사"]] if header_map["제조사"] < len(row) else None)
        major = clean_value(row[header_map["Unit"]] if header_map["Unit"] < len(row) else None)
        minor = clean_value(row[header_map["품목"]] if header_map["품목"] < len(row) else None)
        name = clean_value(row[header_map["모델명/규격"]] if header_map["모델명/규격"] < len(row) else None)
        row_unit = clean_value(row[header_map["단위"]] if "단위" in header_map and header_map["단위"] < len(row) else None)

        qty = parse_int(row[header_map["수량"]] if header_map["수량"] < len(row) else None, default=0)
        if qty <= 0 and major not in (SUMMARY_MAJORS | LABOR_MAJORS):
            continue

        # 일반 부품은 모델명이 필요하지만, 집계/인건비는 name이 비어있을 수 있음
        if not name and major not in (SUMMARY_MAJORS | LABOR_MAJORS):
            continue
        if name.strip() == "-":
            continue

        price = 0
        if "단가" in header_map:
            price = parse_int(row[header_map["단가"]] if header_map["단가"] < len(row) else None, default=0)

        part = find_part(parts_index, maker_name=maker_name, major=major, minor=minor, name=name)
        
        # [신규] 기존 인건비/집계 항목이 부실하면(이름 없음, 가격 0) 업데이트
        if part and major in (SUMMARY_MAJORS | LABOR_MAJORS):
            need_update = False
            update_payload = {}
            
            # 엑셀에서 읽은 유효한 이름 (name이 없으면 minor 사용)
            valid_name = name if name else minor
            
            # 1. 이름 업데이트: DB에 이름이 없는데 엑셀 정보가 있을 때
            # '공백' 텍스트 문제는 clean_value에서 걸러져서 빈 문자열이 됨.
            if not part.get("name") and valid_name:
                 update_payload["name"] = valid_name
                 need_update = True

            # 2. 가격 업데이트: DB 가격이 0인데 엑셀 가격이 있을 때
            if part.get("solo_price", 0) == 0 and int(price) > 0:
                update_payload["solo_price"] = int(price)
                need_update = True

            if need_update:
                put_url = f"{PARTS_CREATE_URL}/{part['resources_id']}/{part['maker_id']}"
                put_json(put_url, update_payload)
                # 다음 로직을 위해 part 정보 갱신
                if "solo_price" in update_payload:
                    part["solo_price"] = update_payload["solo_price"]
                if "name" in update_payload:
                    part["name"] = update_payload["name"]

        if not part:
            # 템플릿 기반으로 마스터(Parts)에 자동 생성 후 1회 재매칭
            created = create_part_from_template_row(
                parts_index,
                {
                    "maker_name": maker_name or (" " if major in (SUMMARY_MAJORS | LABOR_MAJORS) else maker_name),
                    "major_category": major,
                    "minor_category": minor,
                    "name": name,  # labor/summary는 "" 가능
                    "unit": row_unit,
                    "solo_price": price,
                    "ul": parse_cert_bool(row[header_map["UL"]] if "UL" in header_map and header_map["UL"] < len(row) else None),
                    "ce": parse_cert_bool(row[header_map["CE"]] if "CE" in header_map and header_map["CE"] < len(row) else None),
                    "kc": parse_cert_bool(row[header_map["KC"]] if "KC" in header_map and header_map["KC"] < len(row) else None),
                    "certification_etc": clean_value(row[header_map["기타"]] if "기타" in header_map and header_map["기타"] < len(row) else None) or None,
                },
            )
            if created:
                created_parts_count += 1
                part = find_part(parts_index, maker_name=maker_name, major=major, minor=minor, name=name)

        if not part:
            missing.append(
                {
                    "row": row_idx,
                    "maker_name": maker_name,
                    "Unit": major,
                    "품목": minor,
                    "모델명/규격": name,
                    "수량": qty,
                    "단가": price,
                }
            )
            continue

        resources.append(
            {
                "maker_id": part["maker_id"],
                "resources_id": part["resources_id"],
                "solo_price": int(price) if int(price) > 0 else int(part.get("solo_price", 0) or 0),
                "quantity": int(qty),
                # 템플릿 표시값은 마스터(Resources) 분류와 다를 수 있으므로 별도 저장
                "display_major": major or None,
                "display_minor": minor or None,
                "display_model_name": name or None,
                "display_maker_name": maker_name or None,
                "display_unit": row_unit or None,
            }
        )

    return resources, missing, created_parts_count


def main():
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"오류: 파일 경로를 찾을 수 없습니다: {EXCEL_FILE_PATH}")
        return

    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
    print(f"엑셀 로드 완료: 시트 {len(wb.worksheets)}개")

    print("Parts 인덱스 생성 중... (/api/v1/parts)")
    parts_index = fetch_all_parts_index()
    print(
        "Parts 인덱스 생성 완료:"
        f" maker+model={len(parts_index['by_maker_model'])}개"
        f", maker+model(loose)={len(parts_index['by_maker_model_loose'])}개"
        f", major+minor(name='')={len(parts_index['by_major_minor_empty_name'])}개"
    )

    for sheet_index in TEMPLATE_SHEET_INDICES:
        if sheet_index >= len(wb.worksheets):
            print(f"스킵: 시트 인덱스 {sheet_index}가 존재하지 않습니다.")
            continue

        ws = wb.worksheets[sheet_index]
        template_name = ws.title
        machine_name = f"{MACHINE_NAME_PREFIX}{template_name}"

        print(f"\n--- 템플릿 시트 처리: #{sheet_index + 1} '{template_name}' ---")
        try:
            resources, missing, created_parts_count = parse_template_sheet(ws, parts_index)
        except Exception as e:
            print(f"❌ 실패: 시트 파싱 오류: {e}")
            continue

        if missing:
            print(f"⚠️ 매칭 실패 {len(missing)}건 (Parts에 없는 행).")
            for m in missing[:15]:
                print(f"  - row {m['row']}: {m['maker_name']} | {m['Unit']} | {m['품목']} | {m['모델명/규격']} (qty={m['수량']}, price={m['단가']})")
            if len(missing) > 15:
                print("  ... (생략)")
            if STRICT_MODE:
                print("❌ STRICT_MODE=True 이므로 이 시트는 등록하지 않습니다.")
                continue

        if not resources:
            print("스킵: 등록할 리소스가 없습니다(수량>0인 행이 없거나 매칭 실패).")
            continue

        if created_parts_count:
            print(f"🧩 마스터에 자동 추가된 Parts: {created_parts_count}개")

        payload = {
            "name": machine_name,
            "manufacturer": None,
            "client": None,
            "creator": TEMPLATE_CREATOR,
            "description": None,
            "resources": resources,
        }

        existing_id = deduplicate_template_name(machine_name, template_name) if UPSERT_TEMPLATES else None

        if existing_id:
            status, body = put_json(f"{MACHINE_API_URL}{existing_id}", payload)
            action = "갱신"
        else:
            status, body = post_json(MACHINE_API_URL, payload)
            action = "등록"

        if 200 <= status < 300:
            try:
                resp = json.loads(body)
                print(f"✅ {action} 완료: {machine_name} (id={resp.get('id')}, resources={resp.get('resource_count')}, total_price={resp.get('total_price')})")
            except Exception:
                print(f"✅ {action} 완료: {machine_name} (status={status})")
        else:
            preview = body[:300] + ("..." if len(body) > 300 else "")
            print(f"❌ {action} 실패: {machine_name} (status={status}) 응답={preview}")


if __name__ == "__main__":
    main()
