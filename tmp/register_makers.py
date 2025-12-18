# register_makers.py (openpyxl 기반)
import json
import os
from openpyxl import load_workbook
import urllib.request
import urllib.error

# --- Configuration ---
API_URL = "http://localhost:8030/api/v1/maker"
EXCEL_FILE_PATH = "tmp/data.xlsx"
MAKER_SHEET_NAME = 2 # 엑셀 파일의 제조사 데이터 시트 인덱스 (이전 디버깅 결과에 따라 1로 유지)

def clean_value(value):
    """Excel 필드 값을 정리하고 '공백'을 ' '로, 'nan' 등은 빈 문자열로 변환합니다."""
    if value is None:
        return ""

    cleaned = str(value).strip()
    
    # ✅ 수정: '공백'은 띄어쓰기 하나로 변환하여 API에 전송되도록 합니다.
    if cleaned == "공백":
        return " " 
        
    if cleaned in ("nan", "None") or not cleaned:
        return ""
    return cleaned

def get_maker_data_from_excel(excel_file_path, sheet_name):
    """Excel 파일에서 제조사 ID와 Name 목록을 추출합니다."""
    if not os.path.exists(excel_file_path):
        print(f"오류: 파일 경로를 찾을 수 없습니다: {excel_file_path}")
        return []

    try:
        wb = load_workbook(excel_file_path, data_only=True)
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
    except Exception as e:
        print(f"오류: 엑셀 파일/시트({sheet_name}) 읽기 실패. 시트 이름이나 구조를 확인하세요: {e}")
        return []

    required_cols = ['회사명', '회사코드']

    header_row_index = 1  # pandas header=1 과 동일(2번째 줄이 헤더)
    header_cells = [cell.value for cell in ws[header_row_index]]
    header = [str(v).strip() if v is not None else "" for v in header_cells]
    header_map = {name: idx for idx, name in enumerate(header)}

    if not all(col in header_map for col in required_cols):
        print("오류: 엑셀 시트에서 '회사명' 또는 '회사코드' 컬럼을 찾을 수 없습니다.")
        return []

    maker_list = []
    seen_ids = set()

    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        maker_id = clean_value(row[header_map['회사코드']] if header_map['회사코드'] < len(row) else None)
        name = clean_value(row[header_map['회사명']] if header_map['회사명'] < len(row) else None)
        
        # ID와 Name이 완전히 빈 문자열이 아닌 경우에만 등록
        if maker_id != "" and name != "" and maker_id not in seen_ids:
            seen_ids.add(maker_id)
            maker_list.append({
                "id": maker_id,
                "name": name
            })
            
    return maker_list

def post_json(url: str, payload: dict) -> tuple[int, str]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            return resp.getcode(), body
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        return e.code, body

def post_makers_to_api(maker_list, api_url):
    """제조사 데이터 리스트를 API에 POST 요청으로 전송합니다. (이전과 동일)"""
    # ... (전송 로직은 이전과 동일합니다.)
    print(f"\n--- 제조사 API 전송 시작 ({len(maker_list)}개 제조사) ---")
    
    success_count = 0
    failure_count = 0
    
    for maker_data in maker_list:
        maker_name = maker_data['name']
        try:
            status_code, body = post_json(api_url, maker_data)

            if 200 <= status_code < 300:
                success_count += 1
            else:
                if status_code == 409:
                    print(f"⚠️ 이미 존재: {maker_name} (Status: 409)")
                    success_count += 1 
                else:
                    failure_count += 1
                    print(f"❌ 실패: {maker_name} (Status: {status_code}, 응답: {body[:100]}...)")
                
        except urllib.error.URLError as e:
            failure_count += 1
            print(f"🚨 연결 오류: {maker_name} 전송 실패. ({e})")

    print(f"\n--- API 전송 완료 ---")
    print(f"✅ 성공적으로 등록/확인된 제조사: {success_count}개")
    print(f"❌ 등록 실패 제조사: {failure_count}개")

if __name__ == "__main__":
    all_makers_data = get_maker_data_from_excel(EXCEL_FILE_PATH, MAKER_SHEET_NAME)

    if all_makers_data:
        print(f"총 {len(all_makers_data)}개의 고유 제조사 데이터가 추출되었습니다.")
        post_makers_to_api(all_makers_data, API_URL)
    else:
        print("🚨 추출된 제조사 데이터가 없어 API 전송을 시작하지 않습니다.")
