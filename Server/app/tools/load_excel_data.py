# SYNEX+QUOTATION/Server/app/tools/load_excel_data.py
# 엑셀 파일에서 데이터를 읽어 DB에 로드하는 도구
#
# 사용법:
#   cd Server/app
#   DB_MODE=postgresql python tools/load_excel_data.py ../tmp/data.xlsx
#   또는
#   DB_MODE=sqlite python tools/load_excel_data.py ../tmp/data.xlsx

import os
import sys

# 부모 디렉토리를 경로에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy.orm import Session
from database import SessionLocal, engine, Base
from models import Maker, Resources, Certification, Machine, MachineResources


def load_excel_data(excel_path: str):
    """엑셀 파일에서 데이터를 읽어 DB에 로드"""

    print(f"[Load] 엑셀 파일 로드 중: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    db = SessionLocal()

    try:
        # ============================================================
        # 1. Maker 데이터 로드
        # ============================================================
        print("\n[1/4] Maker 데이터 로드 중...")
        maker_sheet = wb['maker']
        maker_map = {}  # 회사명 -> 회사코드 매핑 (정규화된 키)
        maker_map_original = {}  # 원본 회사명 -> 회사코드
        maker_count = 0

        for row_idx in range(2, maker_sheet.max_row + 1):
            name = maker_sheet.cell(row=row_idx, column=1).value
            code = maker_sheet.cell(row=row_idx, column=2).value

            if not name or not code:
                continue

            name = str(name).strip()  # 앞뒤 공백 제거
            # 정규화된 키 (대문자, 공백제거)
            normalized_key = name.upper().replace(" ", "")
            maker_map[normalized_key] = code
            maker_map_original[name] = code

            # DB에 존재하는지 확인
            existing = db.query(Maker).filter(Maker.id == code).first()
            if not existing:
                maker = Maker(id=code, name=name)
                db.add(maker)
                maker_count += 1

        db.commit()
        print(f"  [OK] Maker {maker_count}건 추가 (전체 {len(maker_map)}건)")

        # ============================================================
        # 2. Resources + Certification 데이터 로드
        # ============================================================
        print("\n[2/4] Resources 데이터 로드 중...")
        parts_sheet = wb['parts']
        resources_count = 0
        cert_count = 0

        # Resources ID 생성을 위한 카운터 (maker별)
        id_counters = {}
        # Resources 매핑 저장 (major, minor, name, maker_id) -> res_id
        resources_id_map = {}

        for row_idx in range(2, parts_sheet.max_row + 1):
            major = parts_sheet.cell(row=row_idx, column=1).value  # Unit
            minor = parts_sheet.cell(row=row_idx, column=2).value  # 품목
            name = parts_sheet.cell(row=row_idx, column=3).value   # 모델명/규격
            maker_name = parts_sheet.cell(row=row_idx, column=4).value  # 제조사
            ul = parts_sheet.cell(row=row_idx, column=5).value     # UL
            ce = parts_sheet.cell(row=row_idx, column=6).value     # CE
            kc = parts_sheet.cell(row=row_idx, column=7).value     # KC
            etc = parts_sheet.cell(row=row_idx, column=8).value    # 기타
            unit = parts_sheet.cell(row=row_idx, column=9).value   # 단위
            price = parts_sheet.cell(row=row_idx, column=10).value # 단가

            if not major or not minor or not name:
                continue

            # maker_id 찾기 (정규화된 키로 검색)
            maker_name_str = str(maker_name).strip() if maker_name else ""
            normalized_maker = maker_name_str.upper().replace(" ", "")
            maker_id = maker_map.get(normalized_maker)
            if not maker_id:
                print(f"  [WARN] 제조사 '{maker_name}' 를 찾을 수 없음 (Row {row_idx})")
                continue

            # Resources ID 생성 (maker별 순번)
            if maker_id not in id_counters:
                id_counters[maker_id] = 1
            res_id = f"{id_counters[maker_id]:06d}"
            id_counters[maker_id] += 1

            # 매핑 저장 (나중에 템플릿에서 찾을 때 사용)
            map_key = (str(major)[:50], str(minor)[:50], str(name)[:100], maker_id)
            resources_id_map[map_key] = res_id

            # 가격 처리
            try:
                solo_price = int(price) if price else 0
            except (ValueError, TypeError):
                solo_price = 0

            # DB에 존재하는지 확인
            existing = db.query(Resources).filter(
                Resources.id == res_id,
                Resources.maker_id == maker_id
            ).first()

            if not existing:
                resource = Resources(
                    id=res_id,
                    maker_id=maker_id,
                    major=str(major)[:50],
                    minor=str(minor)[:50],
                    name=str(name)[:100],
                    unit=str(unit)[:10] if unit else 'ea',
                    solo_price=solo_price,
                    display_order=row_idx - 1
                )
                db.add(resource)
                db.flush()  # ID 확정
                resources_count += 1

                # Certification 추가
                cert = Certification(
                    resources_id=res_id,
                    maker_id=maker_id,
                    ul=bool(ul),
                    ce=bool(ce),
                    kc=bool(kc),
                    etc=str(etc) if etc else None
                )
                db.add(cert)
                cert_count += 1

        db.commit()
        print(f"  [OK] Resources {resources_count}건, Certification {cert_count}건 추가")

        # ============================================================
        # 3. Machine (템플릿) 데이터 로드
        # ============================================================
        print("\n[3/4] Machine 템플릿 데이터 로드 중...")

        # 템플릿 시트 목록 (parts, category, maker 제외)
        template_sheets = [s for s in wb.sheetnames if s not in ['parts', 'category', 'maker']]
        machine_count = 0
        machine_map = {}  # 시트명 -> Machine 객체

        for sheet_name in template_sheets:
            # 이미 존재하는지 확인
            existing = db.query(Machine).filter(Machine.name == sheet_name).first()
            if existing:
                machine_map[sheet_name] = existing
                continue

            machine = Machine(
                name=sheet_name,
                manufacturer="SYNEX",
                client=None,
                creator="system",
                price=0,
                description=f"{sheet_name} 템플릿"
            )
            db.add(machine)
            db.flush()
            machine_map[sheet_name] = machine
            machine_count += 1

        db.commit()
        print(f"  [OK] Machine {machine_count}건 추가 (전체 {len(template_sheets)}개 템플릿)")

        # ============================================================
        # 4. MachineResources 데이터 로드
        # ============================================================
        print("\n[4/4] MachineResources 데이터 로드 중...")
        mr_count = 0

        for sheet_name in template_sheets:
            sheet = wb[sheet_name]
            machine = machine_map.get(sheet_name)
            if not machine:
                continue

            # 기존 MachineResources 삭제 (재로드)
            db.query(MachineResources).filter(MachineResources.machine_id == machine.id).delete()

            order_idx = 0
            for row_idx in range(2, sheet.max_row + 1):
                major = sheet.cell(row=row_idx, column=1).value      # Unit
                minor = sheet.cell(row=row_idx, column=2).value      # 품목
                name = sheet.cell(row=row_idx, column=3).value       # 모델명/규격
                maker_name = sheet.cell(row=row_idx, column=4).value # 제조사
                quantity = sheet.cell(row=row_idx, column=9).value   # 수량
                unit = sheet.cell(row=row_idx, column=10).value      # 단위

                if not major or not minor or not name:
                    continue

                # maker_id 찾기 (정규화된 키로 검색)
                maker_name_str = str(maker_name).strip() if maker_name else ""
                normalized_maker = maker_name_str.upper().replace(" ", "")
                m_id = maker_map.get(normalized_maker)
                if not m_id:
                    print(f"  [WARN] 제조사 '{maker_name}' 를 찾을 수 없음 ({sheet_name} Row {row_idx})")
                    continue

                # Resources ID 찾기 (major, minor, name, maker_id 정확히 매칭)
                map_key = (str(major)[:50], str(minor)[:50], str(name)[:100], m_id)
                res_id = resources_id_map.get(map_key)

                if not res_id:
                    print(f"  [WARN] 자재 '{major}/{minor}/{name}' 를 찾을 수 없음 ({sheet_name} Row {row_idx})")
                    continue

                # Resources 객체 조회
                resource = db.query(Resources).filter(
                    Resources.id == res_id,
                    Resources.maker_id == m_id
                ).first()

                if not resource:
                    print(f"  [WARN] DB에서 자재를 찾을 수 없음: {m_id}/{res_id}")
                    continue

                # 수량 처리
                try:
                    qty = int(quantity) if quantity else 1
                except (ValueError, TypeError):
                    qty = 1

                mr = MachineResources(
                    machine_id=machine.id,
                    maker_id=resource.maker_id,
                    resources_id=resource.id,
                    solo_price=resource.solo_price,
                    quantity=qty,
                    order_index=order_idx,
                    display_major=str(major)[:50],
                    display_minor=str(minor)[:50],
                    display_model_name=str(name)[:100],
                    display_maker_name=maker_name,
                    display_unit=str(unit)[:10] if unit else 'ea'
                )
                db.add(mr)
                mr_count += 1
                order_idx += 1

            db.commit()

        print(f"  [OK] MachineResources {mr_count}건 추가")

        # ============================================================
        # 완료
        # ============================================================
        print("\n" + "=" * 50)
        print("[Load] 데이터 로드 완료!")
        print(f"  - Maker: {len(maker_map)}건")
        print(f"  - Resources: {resources_count}건")
        print(f"  - Machine: {len(template_sheets)}건")
        print(f"  - MachineResources: {mr_count}건")
        print("=" * 50)

    except Exception as e:
        db.rollback()
        print(f"\n[ERROR] 에러 발생: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python tools/load_excel_data.py <엑셀파일경로>")
        print("예시: python tools/load_excel_data.py ../tmp/data.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"[ERROR] 파일을 찾을 수 없음: {excel_path}")
        sys.exit(1)

    # 테이블 생성 확인
    from models import *
    Base.metadata.create_all(bind=engine)

    load_excel_data(excel_path)
